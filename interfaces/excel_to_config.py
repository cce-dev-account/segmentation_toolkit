"""
Import SegmentationConfig from Excel Workbook

Parses Excel configuration files and creates SegmentationConfig objects.
Supports validation and error checking.
"""

from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
import sys
from pathlib import Path
import warnings

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from irb_segmentation.config import SegmentationConfig, DataConfig, OutputConfig
from irb_segmentation.params import IRBSegmentationParams


class ConfigFromExcel:
    """Import SegmentationConfig from Excel workbook."""

    @classmethod
    def import_config(cls, excel_path: str) -> SegmentationConfig:
        """
        Import SegmentationConfig from Excel file.

        Args:
            excel_path: Path to Excel configuration file

        Returns:
            SegmentationConfig object

        Raises:
            ValueError: If Excel file is invalid or missing required fields
        """
        # Validate first
        issues = cls.validate_excel(excel_path)
        if issues:
            warnings.warn(f"Excel validation found issues:\n" + "\n".join(f"  - {i}" for i in issues))

        # Load workbook
        wb = load_workbook(excel_path, data_only=True)

        # Parse each section
        metadata = cls._parse_overview(wb)
        data_config = cls._parse_data_config(wb)
        irb_params = cls._parse_irb_params(wb)
        output_config = cls._parse_output_config(wb)

        # Create config
        config = SegmentationConfig(
            data=data_config,
            irb_params=irb_params,
            output=output_config,
            run_validation=metadata.get('run_validation', True),
            verbose=metadata.get('verbose', True),
            name=metadata.get('name'),
            description=metadata.get('description')
        )

        return config

    @classmethod
    def validate_excel(cls, excel_path: str) -> List[str]:
        """
        Validate Excel configuration file.

        Args:
            excel_path: Path to Excel file

        Returns:
            List of validation issues (empty if valid)
        """
        issues = []

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as e:
            issues.append(f"Cannot open Excel file: {e}")
            return issues

        # Check required sheets
        required_sheets = ["Overview", "Data Configuration", "IRB Parameters", "Output Configuration"]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                issues.append(f"Missing required sheet: {sheet}")

        if issues:
            return issues  # Can't validate further without required sheets

        # Validate data config
        try:
            data_issues = cls._validate_data_config(wb)
            issues.extend(data_issues)
        except Exception as e:
            issues.append(f"Error validating Data Configuration: {e}")

        # Validate IRB params
        try:
            irb_issues = cls._validate_irb_params(wb)
            issues.extend(irb_issues)
        except Exception as e:
            issues.append(f"Error validating IRB Parameters: {e}")

        return issues

    @classmethod
    def _parse_overview(cls, wb) -> Dict[str, Any]:
        """Parse Overview sheet."""
        ws = wb["Overview"]

        metadata = {}

        # Parse metadata fields (starting at row 3)
        field_map = {
            "Name:": "name",
            "Description:": "description",
            "Run Validation:": "run_validation",
            "Verbose Output:": "verbose"
        }

        for row in range(3, 10):  # Check first few rows
            label_cell = ws.cell(row, 1).value
            value_cell = ws.cell(row, 2).value

            if label_cell in field_map:
                key = field_map[label_cell]
                if key in ['run_validation', 'verbose']:
                    metadata[key] = cls._parse_bool(value_cell)
                else:
                    metadata[key] = value_cell if value_cell else None

        return metadata

    @classmethod
    def _parse_data_config(cls, wb) -> DataConfig:
        """Parse Data Configuration sheet."""
        ws = wb["Data Configuration"]

        # Parse parameters (row 2 onwards, column B has values)
        data_source = ws.cell(2, 2).value or ''
        data_type = ws.cell(3, 2).value or 'csv'
        sample_size_raw = ws.cell(4, 2).value
        sample_size = int(sample_size_raw) if sample_size_raw and sample_size_raw != '' else None
        use_oot = cls._parse_bool(ws.cell(5, 2).value)
        random_state = int(ws.cell(6, 2).value) if ws.cell(6, 2).value else 42

        # Parse categorical columns (comma-separated)
        categorical_cols_raw = ws.cell(7, 2).value
        if categorical_cols_raw and str(categorical_cols_raw).strip():
            categorical_columns = [col.strip() for col in str(categorical_cols_raw).split(',')]
        else:
            categorical_columns = None

        # Parse target column (row 8)
        target_column_raw = ws.cell(8, 2).value
        target_column = str(target_column_raw).strip() if target_column_raw else None

        return DataConfig(
            source=data_source,
            data_type=data_type,
            sample_size=sample_size,
            use_oot=use_oot,
            random_state=random_state,
            categorical_columns=categorical_columns,
            target_column=target_column
        )

    @classmethod
    def _parse_irb_params(cls, wb) -> IRBSegmentationParams:
        """Parse IRB Parameters sheet."""
        ws = wb["IRB Parameters"]

        # Parse parameters (column C has values)
        # Map row numbers to parameters
        params = {}

        # Tree Structure (rows 2-6)
        params['max_depth'] = int(ws.cell(2, 3).value) if ws.cell(2, 3).value else 3
        params['min_samples_split'] = int(ws.cell(3, 3).value) if ws.cell(3, 3).value else 1000
        params['min_samples_leaf'] = int(ws.cell(4, 3).value) if ws.cell(4, 3).value else 500
        params['criterion'] = ws.cell(5, 3).value or 'gini'
        params['random_state'] = int(ws.cell(6, 3).value) if ws.cell(6, 3).value else 42

        # IRB Requirements (rows 7-9)
        params['min_defaults_per_leaf'] = int(ws.cell(7, 3).value) if ws.cell(7, 3).value else 20
        params['min_default_rate_diff'] = float(ws.cell(8, 3).value) if ws.cell(8, 3).value else 0.001
        params['significance_level'] = float(ws.cell(9, 3).value) if ws.cell(9, 3).value else 0.01

        # Density Controls (rows 10-11)
        params['min_segment_density'] = float(ws.cell(10, 3).value) if ws.cell(10, 3).value else 0.10
        params['max_segment_density'] = float(ws.cell(11, 3).value) if ws.cell(11, 3).value else 0.50

        # Parse business constraints
        params['forced_splits'] = cls._parse_forced_splits(wb)
        params['monotone_constraints'] = cls._parse_monotone_constraints(wb)

        # Parse validation tests
        params['validation_tests'] = cls._parse_validation_tests(wb)

        return IRBSegmentationParams(**params)

    @classmethod
    def _parse_forced_splits(cls, wb) -> Dict:
        """Parse forced splits from Business Constraints sheet."""
        ws = wb["Business Constraints"]

        forced_splits = {}

        # Start reading from row 3 (after header)
        row = 3
        while row < 20:  # Reasonable limit
            feature = ws.cell(row, 1).value
            threshold = ws.cell(row, 2).value
            split_type = ws.cell(row, 3).value

            if not feature or not threshold:
                row += 1
                continue

            # Parse based on type
            if split_type and 'categorical' in split_type.lower():
                # Categorical split - comma-separated values
                forced_splits[feature] = [v.strip() for v in str(threshold).split(',')]
            else:
                # Numeric split
                try:
                    forced_splits[feature] = float(threshold)
                except (ValueError, TypeError):
                    warnings.warn(f"Could not parse threshold for {feature}: {threshold}")

            row += 1

        return forced_splits

    @classmethod
    def _parse_monotone_constraints(cls, wb) -> Dict:
        """Parse monotone constraints from Business Constraints sheet."""
        ws = wb["Business Constraints"]

        monotone_constraints = {}

        # Find monotone constraints section (after forced splits)
        start_row = None
        for row in range(1, 30):
            cell_value = ws.cell(row, 1).value
            if cell_value and 'Monotone Constraints' in str(cell_value):
                start_row = row + 2  # Skip header row
                break

        if not start_row:
            return monotone_constraints

        # Parse monotone constraints
        row = start_row
        while row < start_row + 20:  # Reasonable limit
            feature = ws.cell(row, 1).value
            direction = ws.cell(row, 2).value

            if not feature or not direction:
                row += 1
                continue

            # Map direction to integer
            direction_map = {
                'Increasing': 1,
                'Decreasing': -1,
                'None': 0,
                '1': 1,
                '-1': -1,
                '0': 0
            }

            if str(direction) in direction_map:
                monotone_constraints[feature] = direction_map[str(direction)]
            else:
                warnings.warn(f"Unknown direction for {feature}: {direction}")

            row += 1

        return monotone_constraints

    @classmethod
    def _parse_validation_tests(cls, wb) -> List[str]:
        """Parse validation tests from Validation Tests sheet."""
        if "Validation Tests" not in wb.sheetnames:
            return ['chi_squared', 'psi', 'binomial']  # Defaults

        ws = wb["Validation Tests"]

        tests = []

        # Start from row 2 (after header)
        for row in range(2, 10):  # Check a few rows
            test_name = ws.cell(row, 1).value
            include = ws.cell(row, 2).value

            if test_name and cls._parse_bool(include):
                tests.append(test_name)

        return tests if tests else ['chi_squared', 'psi', 'binomial']

    @classmethod
    def _parse_output_config(cls, wb) -> OutputConfig:
        """Parse Output Configuration sheet."""
        ws = wb["Output Configuration"]

        # Parse parameters (row 2 onwards, column B has values)
        output_dir = ws.cell(2, 2).value or './output'
        output_formats_raw = ws.cell(3, 2).value or 'json,html'
        output_formats = [f.strip() for f in str(output_formats_raw).split(',')]

        report_name = ws.cell(4, 2).value
        if report_name and 'auto' in str(report_name).lower():
            report_name = None

        template_name = ws.cell(5, 2).value
        if template_name and 'auto' in str(template_name).lower():
            template_name = None

        dashboard_name = ws.cell(6, 2).value
        if dashboard_name and 'auto' in str(dashboard_name).lower():
            dashboard_name = None

        create_dashboard = cls._parse_bool(ws.cell(7, 2).value)
        create_excel_template = cls._parse_bool(ws.cell(8, 2).value)
        extract_rules = cls._parse_bool(ws.cell(9, 2).value)

        return OutputConfig(
            output_dir=output_dir,
            output_formats=output_formats,
            report_name=report_name,
            template_name=template_name,
            dashboard_name=dashboard_name,
            create_dashboard=create_dashboard,
            create_excel_template=create_excel_template,
            extract_rules=extract_rules
        )

    @classmethod
    def _validate_data_config(cls, wb) -> List[str]:
        """Validate data configuration."""
        issues = []

        ws = wb["Data Configuration"]

        data_source = ws.cell(2, 2).value
        if not data_source:
            issues.append("Data source is required")

        data_type = ws.cell(3, 2).value
        valid_types = ['csv', 'german_credit', 'lending_club', 'taiwan_credit', 'home_credit']
        if data_type and data_type not in valid_types:
            issues.append(f"Invalid data_type: {data_type}. Must be one of {valid_types}")

        return issues

    @classmethod
    def _validate_irb_params(cls, wb) -> List[str]:
        """Validate IRB parameters."""
        issues = []

        ws = wb["IRB Parameters"]

        # Validate max_depth
        max_depth = ws.cell(2, 3).value
        if max_depth and (max_depth < 1 or max_depth > 10):
            issues.append(f"max_depth ({max_depth}) should be between 1 and 10")

        # Validate min_defaults_per_leaf
        min_defaults = ws.cell(7, 3).value
        if min_defaults and min_defaults < 20:
            issues.append(f"Warning: min_defaults_per_leaf ({min_defaults}) is below Basel recommended minimum of 20")

        # Validate density constraints
        min_density = ws.cell(10, 3).value
        max_density = ws.cell(11, 3).value
        if min_density and max_density and min_density >= max_density:
            issues.append(f"min_segment_density ({min_density}) must be less than max_segment_density ({max_density})")

        return issues

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        """Parse boolean value from Excel."""
        if value is None:
            return False

        if isinstance(value, bool):
            return value

        value_str = str(value).strip().lower()
        return value_str in ['yes', 'true', '1', 'y']


if __name__ == '__main__':
    # Test import
    config = ConfigFromExcel.import_config('test_config.xlsx')
    print("Import successful!")
    print(config.summary())
