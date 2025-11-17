"""
Convert Excel/CSV Templates to JSON Modification File

Reads the Excel workbook or CSV files created by create_excel_template.py
and generates a modification.json file for apply_modifications.py
"""

import csv
import json
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def read_segment_actions(filepath):
    """Read segment actions CSV and extract merge decisions."""
    merges = []
    notes = []

    with open(filepath, 'r', encoding='utf-8') as f:
        # Read all lines
        lines = f.readlines()

        # Find the data header row (contains Segment_ID)
        header_idx = None
        for i, line in enumerate(lines):
            if 'Segment_ID' in line and 'Action' in line:
                header_idx = i
                break

        if header_idx is None:
            print("Warning: Could not find data header in segment actions file")
            return merges, notes

        # Parse from header onwards
        import io
        data_lines = lines[header_idx:]
        reader = csv.DictReader(io.StringIO(''.join(data_lines)))

        for row in reader:
            # Skip empty or header rows
            if not row.get('Segment_ID') or row['Segment_ID'].startswith('Valid'):
                continue

            try:
                seg_id = int(row['Segment_ID'])
            except (ValueError, TypeError):
                continue

            action = (row.get('Action') or '').strip().upper()
            merge_into = (row.get('Merge_Into') or '').strip()
            note = (row.get('Notes') or '').strip()

            if action == 'MERGE' and merge_into:
                try:
                    target_id = int(merge_into)
                    merges.append([seg_id, target_id])

                    if note:
                        notes.append(f"Segment {seg_id} → {target_id}: {note}")
                except (ValueError, TypeError):
                    print(f"Warning: Invalid merge target '{merge_into}' for segment {seg_id}")

    return merges, notes


def read_thresholds(filepath):
    """Read threshold editor CSV and extract forced splits."""
    forced_splits = {}
    notes = []

    with open(filepath, 'r', encoding='utf-8') as f:
        import io
        lines = f.readlines()

        # Find header row and end of data section
        header_idx = None
        end_idx = None
        for i, line in enumerate(lines):
            if 'Feature_Name' in line and 'Threshold_Value' in line:
                header_idx = i
            if header_idx is not None and 'CURRENT THRESHOLDS' in line:
                end_idx = i
                break

        if header_idx is None:
            print("Warning: Could not find data header in thresholds file")
            return forced_splits, notes

        # Only read data section (between header and CURRENT THRESHOLDS marker)
        if end_idx:
            data_lines = lines[header_idx:end_idx]
        else:
            data_lines = lines[header_idx:]

        reader = csv.DictReader(io.StringIO(''.join(data_lines)))

        for row in reader:
            feature = (row.get('Feature_Name') or '').strip()
            threshold_str = (row.get('Threshold_Value') or '').strip()
            reason = (row.get('Business_Reason') or '').strip()

            # Skip empty rows
            if not feature or not threshold_str:
                continue

            # Skip reference/header/description rows
            if (feature.startswith('CURRENT') or feature.startswith('AVAILABLE') or
                feature in ['Feature', 'Feature_Name'] or
                threshold_str in ['Threshold', 'Existing_Thresholds'] or
                ',' in threshold_str or  # Skip multi-value reference rows
                threshold_str.endswith(')')):  # Skip description rows
                continue

            try:
                threshold = float(threshold_str.replace(',', ''))
                forced_splits[feature] = threshold

                if reason:
                    notes.append(f"{feature} @ {threshold}: {reason}")
            except (ValueError, TypeError):
                print(f"Warning: Invalid threshold value '{threshold_str}' for feature '{feature}'")

    return forced_splits, notes


def read_parameters(filepath):
    """Read parameters CSV and extract changes."""
    parameters = {}
    notes = []

    with open(filepath, 'r', encoding='utf-8') as f:
        import io
        lines = f.readlines()

        # Find header row
        header_idx = None
        for i, line in enumerate(lines):
            if 'Parameter' in line and 'New_Value' in line:
                header_idx = i
                break

        if header_idx is None:
            print("Warning: Could not find data header in parameters file")
            return parameters, notes

        data_lines = lines[header_idx:]
        reader = csv.DictReader(io.StringIO(''.join(data_lines)))

        for row in reader:
            param_name = (row.get('Parameter') or '').strip()
            new_value_str = (row.get('New_Value') or '').strip()
            current_value_str = (row.get('Current_Value') or '').strip()
            reason = (row.get('Change_Reason') or '').strip()

            if not param_name or not new_value_str:
                continue

            # Skip header rows
            if param_name in ['Parameter', 'Change'] or param_name.startswith('PARAMETER'):
                continue

            try:
                # Determine if it's float or int
                if '.' in new_value_str:
                    new_value = float(new_value_str)
                else:
                    new_value = int(new_value_str)

                parameters[param_name] = new_value

                if reason:
                    notes.append(f"{param_name}: {current_value_str} → {new_value} ({reason})")
            except (ValueError, TypeError):
                print(f"Warning: Invalid parameter value '{new_value_str}' for '{param_name}'")

    return parameters, notes


def csv_to_json(
    segment_actions_file: str = "modification_template_segment_actions.csv",
    thresholds_file: str = "modification_template_thresholds.csv",
    parameters_file: str = "modification_template_parameters.csv",
    output_file: str = "modification.json"
):
    """Convert CSV templates to JSON modification file."""

    print("\n" + "=" * 80)
    print("CONVERTING CSV TEMPLATES TO JSON")
    print("=" * 80)

    # Resolve paths
    templates_dir = Path(__file__).parent / 'templates'
    output_path = Path(__file__).parent.parent / output_file

    segment_path = templates_dir / segment_actions_file
    threshold_path = templates_dir / thresholds_file
    param_path = templates_dir / parameters_file

    # Check files exist
    missing = []
    if not segment_path.exists():
        missing.append(segment_path.name)
    if not threshold_path.exists():
        missing.append(threshold_path.name)
    if not param_path.exists():
        missing.append(param_path.name)

    if missing:
        print(f"\nError: Missing template files: {', '.join(missing)}")
        print(f"Run: python interfaces/create_excel_template.py")
        return False

    print(f"\nReading templates from: {templates_dir}")

    # Read all three files
    print("\n1. Reading segment actions...")
    merges, merge_notes = read_segment_actions(segment_path)
    print(f"   Found {len(merges)} segment merge(s)")

    print("\n2. Reading thresholds...")
    forced_splits, threshold_notes = read_thresholds(threshold_path)
    print(f"   Found {len(forced_splits)} forced split(s)")

    print("\n3. Reading parameters...")
    parameters, param_notes = read_parameters(param_path)
    print(f"   Found {len(parameters)} parameter change(s)")

    # Build JSON structure
    modification = {
        "metadata": {
            "instructions": "Generated from CSV templates by excel_to_json.py",
            "source_files": [
                segment_actions_file,
                thresholds_file,
                parameters_file
            ],
            "modification_notes": merge_notes + threshold_notes + param_notes
        },
        "modifications": {
            "merge_segments": {
                "description": "Segment pairs to merge",
                "value": merges
            },
            "forced_splits": {
                "description": "Forced split thresholds",
                "value": forced_splits
            },
            "parameter_changes": parameters
        }
    }

    # Validate
    print("\n" + "-" * 80)
    print("VALIDATION")
    print("-" * 80)

    issues = []

    # Check for duplicate merges
    merge_sources = [m[0] for m in merges]
    if len(merge_sources) != len(set(merge_sources)):
        issues.append("Duplicate merge sources detected")

    # Check for self-merges
    for seg1, seg2 in merges:
        if seg1 == seg2:
            issues.append(f"Cannot merge segment {seg1} into itself")

    # Check parameter ranges
    if 'max_depth' in parameters:
        if not 2 <= parameters['max_depth'] <= 10:
            issues.append(f"max_depth {parameters['max_depth']} outside recommended range 2-10")

    if 'min_samples_leaf' in parameters:
        if parameters['min_samples_leaf'] < 100:
            issues.append(f"min_samples_leaf {parameters['min_samples_leaf']} too low (minimum 100)")

    if 'min_segment_density' in parameters and 'max_segment_density' in parameters:
        if parameters['min_segment_density'] >= parameters['max_segment_density']:
            issues.append("min_segment_density must be less than max_segment_density")

    if issues:
        print("[WARN] Validation warnings:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("[OK] All validations passed")

    # Save JSON
    with open(output_path, 'w') as f:
        json.dump(modification, f, indent=2)

    print("\n" + "=" * 80)
    print("CONVERSION COMPLETE")
    print("=" * 80)
    print(f"\nOutput file: {output_path}")

    print("\nSummary:")
    print(f"  Segment merges: {len(merges)}")
    if merges:
        for seg1, seg2 in merges:
            print(f"    - Segment {seg1} -> {seg2}")

    print(f"  Forced splits: {len(forced_splits)}")
    if forced_splits:
        for feature, threshold in forced_splits.items():
            print(f"    - {feature} @ {threshold}")

    print(f"  Parameter changes: {len(parameters)}")
    if parameters:
        for param, value in parameters.items():
            print(f"    - {param} = {value}")

    print("\nNext step:")
    print(f"  python apply_modifications.py {output_file}")

    return True


def excel_workbook_to_json(
    workbook_file: str = "modification_template.xlsx",
    output_file: str = "modification.json"
):
    """Convert Excel workbook to JSON modification file."""

    if not EXCEL_AVAILABLE:
        print("\nError: openpyxl not available. Install with: pip install openpyxl")
        print("Or use CSV format instead.")
        return False

    print("\n" + "=" * 80)
    print("CONVERTING EXCEL WORKBOOK TO JSON")
    print("=" * 80)

    workbook_path = Path(__file__).parent.parent / workbook_file
    output_path = Path(__file__).parent.parent / output_file

    if not workbook_path.exists():
        print(f"\nError: Workbook not found: {workbook_path}")
        print(f"Run: python interfaces/create_excel_template.py")
        return False

    # Load workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Read Segment Actions worksheet
    print("\n1. Reading Segment Actions worksheet...")
    merges = []
    merge_notes = []

    ws_actions = wb["Segment Actions"]
    for row in ws_actions.iter_rows(min_row=9, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        seg_id = row[0]
        action = (row[6] or '').strip().upper() if len(row) > 6 else ''
        merge_into = (row[7] or '').strip() if len(row) > 7 else ''
        notes = (row[8] or '').strip() if len(row) > 8 else ''

        if action == 'MERGE' and merge_into:
            try:
                target_id = int(merge_into)
                merges.append([int(seg_id), target_id])
                if notes:
                    merge_notes.append(f"Segment {seg_id} -> {target_id}: {notes}")
            except (ValueError, TypeError):
                print(f"Warning: Invalid merge target '{merge_into}' for segment {seg_id}")

    print(f"   Found {len(merges)} segment merge(s)")

    # Read threshold changes - try Segment Rules first, fall back to feature sheets
    print("\n2. Reading threshold changes...")
    forced_splits = {}
    threshold_notes = []

    # Check if Segment Rules worksheet exists (new format)
    if "Segment Rules" in wb.sheetnames:
        print("   Using 'Segment Rules' worksheet (segment-centric view)")
        ws_seg_rules = wb["Segment Rules"]

        # Read from row 9 onwards, looking for threshold edits
        for row in ws_seg_rules.iter_rows(min_row=9, values_only=True):
            if not row or len(row) < 4:
                continue

            # Columns: Feature | Condition | Current Value | New Value | Format | Business Reason
            feature = (row[0] or '').strip()
            condition = (row[1] or '').strip()
            current_val = row[2]
            new_val = row[3]
            fmt = (row[4] or '').strip() if len(row) > 4 else ''
            reason = (row[5] or '').strip() if len(row) > 5 else ''

            # Skip if not a data row (headers, segment headers, etc.)
            if not feature or feature.startswith('SEGMENT') or feature in ['Feature', 'Path']:
                continue

            # Only process if user entered a new value
            if new_val is not None and new_val != '':
                try:
                    # Parse the value (remove formatting)
                    if isinstance(new_val, (int, float)):
                        threshold = float(new_val)
                    else:
                        new_val_str = str(new_val).strip()
                        # Remove formatting
                        new_val_str = new_val_str.replace('$', '').replace('%', '').replace(',', '').strip()
                        if new_val_str:
                            threshold = float(new_val_str)
                        else:
                            continue

                    # Store the threshold
                    if feature not in forced_splits:
                        forced_splits[feature] = []

                    # Check if this threshold is already in the list
                    if threshold not in forced_splits[feature]:
                        forced_splits[feature].append(threshold)

                    if reason:
                        threshold_notes.append(f"{feature} {condition} {threshold}: {reason}")

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid threshold value '{new_val}' for {feature}: {e}")

        # Convert lists to single values or keep as sorted lists
        final_splits = {}
        for feature, thresholds in forced_splits.items():
            if len(thresholds) == 1:
                final_splits[feature] = thresholds[0]
            else:
                final_splits[feature] = sorted(set(thresholds))

        forced_splits = final_splits
        print(f"   Found {sum(len(v) if isinstance(v, list) else 1 for v in forced_splits.values())} threshold change(s) across {len(forced_splits)} feature(s)")

    else:
        # Fall back to reading individual feature threshold worksheets (old format)
        print("   Using individual feature worksheets (feature-centric view)")

    # List of expected feature worksheets
    feature_sheets = ['int_rate', 'annual_inc', 'fico_range_high', 'fico_range_low',
                     'dti', 'inq_last_6mths', 'installment', 'loan_amnt']

    for feature in feature_sheets:
        if feature not in wb.sheetnames:
            continue  # Skip if sheet doesn't exist

        ws_feature = wb[feature]

        # Read from row 9 onwards (after headers)
        for row in ws_feature.iter_rows(min_row=9, values_only=True):
            if not row or len(row) < 3:
                continue

            # Columns: Order | Current_Value | New_Value | Used_By_Segments | Business_Reason
            current_val = row[1]  # Current value (formatted)
            new_val = row[2]      # New value (user edited)
            reason = (row[4] or '').strip() if len(row) > 4 else ''

            # Only process if user entered a new value
            if new_val is not None and new_val != '':
                try:
                    # Parse the value (remove formatting like $, %, etc.)
                    if isinstance(new_val, (int, float)):
                        threshold = float(new_val)
                    else:
                        new_val_str = str(new_val).strip()
                        # Remove currency symbols and percentage signs
                        new_val_str = new_val_str.replace('$', '').replace('%', '').replace(',', '').strip()
                        if new_val_str:
                            threshold = float(new_val_str)
                        else:
                            continue

                    # Store the threshold for this feature
                    if feature not in forced_splits:
                        forced_splits[feature] = []

                    forced_splits[feature].append(threshold)

                    if reason:
                        threshold_notes.append(f"{feature} @ {threshold}: {reason}")

                except (ValueError, TypeError) as e:
                    print(f"Warning: Invalid threshold value '{new_val}' for feature '{feature}': {e}")

    # Convert lists to single values or keep as lists
    final_splits = {}
    for feature, thresholds in forced_splits.items():
        if isinstance(thresholds, list) and len(thresholds) == 1:
            final_splits[feature] = thresholds[0]
        elif isinstance(thresholds, list):
            # Multiple thresholds - keep as sorted list
            final_splits[feature] = sorted(set(thresholds))
        else:
            # Already a single value
            final_splits[feature] = thresholds

    forced_splits = final_splits
    print(f"   Found {sum(len(v) if isinstance(v, list) else 1 for v in forced_splits.values())} threshold change(s) across {len(forced_splits)} feature(s)")

    # Read Model Parameters worksheet
    print("\n3. Reading Model Parameters worksheet...")
    parameters = {}
    param_notes = []

    ws_params = wb["Model Parameters"]
    for row in ws_params.iter_rows(min_row=8, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        param_name = (row[0] or '').strip()
        current_value = row[1]
        new_value = row[2]
        reason = (row[5] or '').strip() if len(row) > 5 else ''

        if new_value is not None and new_value != '':
            try:
                # Determine if it's float or int
                if isinstance(new_value, (int, float)):
                    param_value = new_value
                else:
                    new_value_str = str(new_value)
                    if '.' in new_value_str:
                        param_value = float(new_value_str)
                    else:
                        param_value = int(new_value_str)

                parameters[param_name] = param_value
                if reason:
                    param_notes.append(f"{param_name}: {current_value} -> {param_value} ({reason})")
            except (ValueError, TypeError):
                print(f"Warning: Invalid parameter value '{new_value}' for '{param_name}'")

    print(f"   Found {len(parameters)} parameter change(s)")

    # Build JSON structure
    modification = {
        "metadata": {
            "instructions": "Generated from Excel workbook by excel_to_json.py",
            "source_files": [workbook_file],
            "modification_notes": merge_notes + threshold_notes + param_notes
        },
        "modifications": {
            "merge_segments": {
                "description": "Segment pairs to merge",
                "value": merges
            },
            "forced_splits": {
                "description": "Forced split thresholds",
                "value": forced_splits
            },
            "parameter_changes": parameters
        }
    }

    # Validate
    print("\n" + "-" * 80)
    print("VALIDATION")
    print("-" * 80)

    issues = []

    # Check for duplicate merges
    merge_sources = [m[0] for m in merges]
    if len(merge_sources) != len(set(merge_sources)):
        issues.append("Duplicate merge sources detected")

    # Check for self-merges
    for seg1, seg2 in merges:
        if seg1 == seg2:
            issues.append(f"Cannot merge segment {seg1} into itself")

    # Check parameter ranges
    if 'max_depth' in parameters:
        if not 2 <= parameters['max_depth'] <= 10:
            issues.append(f"max_depth {parameters['max_depth']} outside recommended range 2-10")

    if 'min_samples_leaf' in parameters:
        if parameters['min_samples_leaf'] < 100:
            issues.append(f"min_samples_leaf {parameters['min_samples_leaf']} too low (minimum 100)")

    if 'min_segment_density' in parameters and 'max_segment_density' in parameters:
        if parameters['min_segment_density'] >= parameters['max_segment_density']:
            issues.append("min_segment_density must be less than max_segment_density")

    if issues:
        print("[WARN] Validation warnings:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("[OK] All validations passed")

    # Save JSON
    with open(output_path, 'w') as f:
        json.dump(modification, f, indent=2)

    print("\n" + "=" * 80)
    print("CONVERSION COMPLETE")
    print("=" * 80)
    print(f"\nOutput file: {output_path}")

    print("\nSummary:")
    print(f"  Segment merges: {len(merges)}")
    if merges:
        for seg1, seg2 in merges:
            print(f"    - Segment {seg1} -> {seg2}")

    print(f"  Forced splits: {len(forced_splits)}")
    if forced_splits:
        for feature, threshold in forced_splits.items():
            print(f"    - {feature} @ {threshold}")

    print(f"  Parameter changes: {len(parameters)}")
    if parameters:
        for param, value in parameters.items():
            print(f"    - {param} = {value}")

    print("\nNext step:")
    print(f"  python apply_modifications.py {output_file}")

    return True


def main():
    """Main entry point."""
    # Check if Excel workbook exists first
    workbook_path = Path(__file__).parent.parent / "modification_template.xlsx"

    if workbook_path.exists() and EXCEL_AVAILABLE:
        print("\nDetected Excel workbook format")
        success = excel_workbook_to_json()
    else:
        print("\nUsing CSV format")
        success = csv_to_json()

    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
