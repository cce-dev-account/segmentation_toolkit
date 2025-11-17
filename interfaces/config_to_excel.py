"""
Export SegmentationConfig to Excel Workbook

Creates user-friendly Excel interface for configuration editing with:
- Data validation (dropdowns, ranges)
- Conditional formatting (Basel compliance indicators)
- Help text and documentation
- Multiple template types (simple, advanced, comparison)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from typing import Dict, List, Any
import sys
from pathlib import Path

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from irb_segmentation.config import SegmentationConfig


class ConfigToExcel:
    """Export SegmentationConfig to Excel workbook."""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Yellow
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Red

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FONT = Font(bold=True, size=10)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @classmethod
    def export_config(cls, config: SegmentationConfig, output_path: str, template_type: str = 'standard'):
        """
        Export SegmentationConfig to Excel workbook.

        Args:
            config: Configuration to export
            output_path: Path to save Excel file
            template_type: 'simple', 'standard', or 'advanced'
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        cls._create_overview_sheet(wb, config)
        cls._create_data_sheet(wb, config)
        cls._create_irb_params_sheet(wb, config, template_type)
        cls._create_business_constraints_sheet(wb, config)
        cls._create_output_sheet(wb, config)
        cls._create_validation_sheet(wb, config)

        if template_type in ['standard', 'advanced']:
            cls._create_help_sheet(wb)

        # Save workbook
        wb.save(output_path)
        print(f"Configuration exported to Excel: {output_path}")

    @classmethod
    def create_template(cls, output_path: str, template_type: str = 'simple'):
        """
        Create blank Excel template.

        Args:
            output_path: Path to save template
            template_type: 'simple', 'standard', or 'advanced'
        """
        # Create empty config with defaults
        from irb_segmentation import create_default_config
        config = create_default_config(
            data_source='',
            data_type='csv',
            output_dir='./output'
        )

        cls.export_config(config, output_path, template_type)

    @classmethod
    def _create_overview_sheet(cls, wb: Workbook, config: SegmentationConfig):
        """Create Overview sheet."""
        ws = wb.create_sheet("Overview")

        # Title
        ws['A1'] = "IRB Segmentation Configuration"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Metadata
        row = 3
        fields = [
            ("Name:", config.name or ''),
            ("Description:", config.description or ''),
            ("Run Validation:", "Yes" if config.run_validation else "No"),
            ("Verbose Output:", "Yes" if config.verbose else "No"),
        ]

        for label, value in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Instructions
        row += 2
        ws[f'A{row}'] = "Instructions:"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1

        instructions = [
            "1. Review and edit parameters in each sheet",
            "2. Use dropdowns where provided",
            "3. See 'Help' sheet for parameter descriptions",
            "4. Green cells = Basel compliant, Yellow = caution, Red = non-compliant",
            "5. Save file and import using: ConfigBuilder.from_excel('filename.xlsx')"
        ]

        for instruction in instructions:
            ws[f'A{row}'] = instruction
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50

    @classmethod
    def _create_data_sheet(cls, wb: Workbook, config: SegmentationConfig):
        """Create Data Configuration sheet."""
        ws = wb.create_sheet("Data Configuration")

        # Header
        headers = ["Parameter", "Value", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # Parameters
        categorical_cols_str = ', '.join(config.data.categorical_columns) if config.data.categorical_columns else ''
        params = [
            ("Data Source", config.data.source, "Path to CSV file or loader name (german_credit, lending_club, etc.)"),
            ("Data Type", config.data.data_type, "Type of data source"),
            ("Sample Size", config.data.sample_size or '', "Leave blank for full dataset, or specify number"),
            ("Use Out-of-Time Split", "Yes" if config.data.use_oot else "No", "Use temporal validation split if available"),
            ("Random State", config.data.random_state, "Random seed for reproducibility"),
            ("Categorical Columns", categorical_cols_str, "Comma-separated list of categorical column names (e.g., loan_purpose, grade)"),
            ("Target Column", config.data.target_column or '', "Specify target column name for CSV files (leave blank for auto-detection)"),
        ]

        for row, (param, value, desc) in enumerate(params, 2):
            ws.cell(row, 1, param).font = Font(bold=True)
            ws.cell(row, 2, value)
            ws.cell(row, 3, desc)

        # Add data validation for data_type
        data_type_validation = DataValidation(
            type="list",
            formula1='"csv,german_credit,lending_club,taiwan_credit,home_credit"',
            allow_blank=False
        )
        data_type_validation.add(ws['B3'])
        ws.add_data_validation(data_type_validation)

        # Add data validation for boolean fields
        bool_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        bool_validation.add(ws['B5'])
        ws.add_data_validation(bool_validation)

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60

        # Apply borders
        for row in range(1, len(params) + 2):
            for col in range(1, 4):
                ws.cell(row, col).border = cls.BORDER

    @classmethod
    def _create_irb_params_sheet(cls, wb: Workbook, config: SegmentationConfig, template_type: str):
        """Create IRB Parameters sheet."""
        ws = wb.create_sheet("IRB Parameters")

        # Header
        headers = ["Category", "Parameter", "Value", "Min", "Max", "Basel Rec", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # Parameters organized by category
        irb = config.irb_params

        params = [
            # Tree Structure
            ("Tree Structure", "Max Depth", irb.max_depth, 1, 10, "3-5", "Maximum depth of decision tree"),
            ("", "Min Samples Split", irb.min_samples_split, 2, None, "-", "Minimum samples required to split a node"),
            ("", "Min Samples Leaf", irb.min_samples_leaf, 1, None, "-", "Minimum samples required in a leaf node"),
            ("", "Criterion", irb.criterion, None, None, "gini", "Split quality measure"),
            ("", "Random State", irb.random_state, None, None, "42", "Random seed for reproducibility"),

            # IRB Requirements
            ("IRB Requirements", "Min Defaults Per Leaf", irb.min_defaults_per_leaf, 1, None, "≥20", "Minimum default events per segment (Basel regulatory)"),
            ("", "Min Default Rate Diff", irb.min_default_rate_diff, 0.0, 1.0, "0.001", "Minimum PD difference between segments"),
            ("", "Significance Level", irb.significance_level, 0.0, 1.0, "0.01", "Statistical significance level for tests"),

            # Density Controls
            ("Density Controls", "Min Segment Density", irb.min_segment_density, 0.0, 1.0, "0.10", "Minimum proportion of observations per segment"),
            ("", "Max Segment Density", irb.max_segment_density, 0.0, 1.0, "0.50", "Maximum proportion of observations per segment"),
        ]

        row = 2
        for category, param, value, min_val, max_val, basel_rec, desc in params:
            ws.cell(row, 1, category).font = cls.SECTION_FONT if category else Font()
            ws.cell(row, 2, param).font = Font(bold=True)
            ws.cell(row, 3, value)
            ws.cell(row, 4, min_val if min_val is not None else '-')
            ws.cell(row, 5, max_val if max_val is not None else '-')
            ws.cell(row, 6, basel_rec)
            ws.cell(row, 7, desc)

            # Apply conditional formatting based on Basel recommendations
            if param == "Min Defaults Per Leaf":
                if value >= 20:
                    ws.cell(row, 3).fill = cls.COMPLIANT_FILL
                elif value >= 10:
                    ws.cell(row, 3).fill = cls.WARNING_FILL
                else:
                    ws.cell(row, 3).fill = cls.ERROR_FILL

            row += 1

        # Add data validation for Criterion
        criterion_validation = DataValidation(
            type="list",
            formula1='"gini,entropy"',
            allow_blank=False
        )
        criterion_validation.add(ws['C5'])
        ws.add_data_validation(criterion_validation)

        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50

        # Apply borders
        for r in range(1, row):
            for col in range(1, 8):
                ws.cell(r, col).border = cls.BORDER

    @classmethod
    def _create_business_constraints_sheet(cls, wb: Workbook, config: SegmentationConfig):
        """Create Business Constraints sheet."""
        ws = wb.create_sheet("Business Constraints")

        # Forced Splits section
        ws['A1'] = "Forced Splits"
        ws['A1'].font = cls.SECTION_FONT
        ws['A1'].fill = cls.SECTION_FILL
        ws.merge_cells('A1:D1')

        ws['A2'] = "Feature Name"
        ws['B2'] = "Threshold/Values"
        ws['C2'] = "Type"
        ws['D2'] = "Reason"

        for col in range(1, 5):
            ws.cell(2, col).font = cls.HEADER_FONT
            ws.cell(2, col).fill = cls.HEADER_FILL

        # Add forced splits
        row = 3
        for feature, value in config.irb_params.forced_splits.items():
            ws.cell(row, 1, feature)
            if isinstance(value, list):
                ws.cell(row, 2, ','.join(map(str, value)))
                ws.cell(row, 3, "Categorical")
            else:
                ws.cell(row, 2, value)
                ws.cell(row, 3, "Numeric")
            ws.cell(row, 4, '')  # Reason - user can fill
            row += 1

        # Add empty rows for user to add more
        for _ in range(3):
            for col in range(1, 5):
                ws.cell(row, col, '')
            row += 1

        # Monotone Constraints section
        row += 2
        ws[f'A{row}'] = "Monotone Constraints"
        ws[f'A{row}'].font = cls.SECTION_FONT
        ws[f'A{row}'].fill = cls.SECTION_FILL
        ws.merge_cells(f'A{row}:C{row}')

        row += 1
        ws.cell(row, 1, "Feature Name").font = cls.HEADER_FONT
        ws.cell(row, 2, "Direction").font = cls.HEADER_FONT
        ws.cell(row, 3, "Description").font = cls.HEADER_FONT

        for col in range(1, 4):
            ws.cell(row, col).fill = cls.HEADER_FILL

        # Add monotone constraints
        row += 1
        direction_map = {1: "Increasing", -1: "Decreasing", 0: "None"}
        for feature, direction in config.irb_params.monotone_constraints.items():
            ws.cell(row, 1, feature)
            ws.cell(row, 2, direction_map.get(direction, direction))
            ws.cell(row, 3, '')  # Description - user can fill
            row += 1

        # Add empty rows for user to add more
        for _ in range(3):
            for col in range(1, 4):
                ws.cell(row, col, '')
            row += 1

        # Add data validation for Direction
        direction_validation = DataValidation(
            type="list",
            formula1='"Increasing,Decreasing,None"',
            allow_blank=True
        )
        # Apply to reasonable range
        direction_validation.add(f'B{row-3}:B{row+10}')
        ws.add_data_validation(direction_validation)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

    @classmethod
    def _create_output_sheet(cls, wb: Workbook, config: SegmentationConfig):
        """Create Output Configuration sheet."""
        ws = wb.create_sheet("Output Configuration")

        # Header
        headers = ["Parameter", "Value", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL

        # Parameters
        params = [
            ("Output Directory", config.output.output_dir, "Directory to save all outputs"),
            ("Output Formats", ','.join(config.output.output_formats), "Comma-separated: json, excel, html, yaml"),
            ("Report Name", config.output.report_name or 'Auto-generated', "Custom report name (optional)"),
            ("Template Name", config.output.template_name or 'Auto-generated', "Custom template name (optional)"),
            ("Dashboard Name", config.output.dashboard_name or 'Auto-generated', "Custom dashboard name (optional)"),
            ("Create Dashboard", "Yes" if config.output.create_dashboard else "No", "Generate HTML dashboard"),
            ("Create Excel Template", "Yes" if config.output.create_excel_template else "No", "Generate Excel modification template"),
            ("Extract Rules", "Yes" if config.output.extract_rules else "No", "Extract segment decision rules"),
        ]

        for row, (param, value, desc) in enumerate(params, 2):
            ws.cell(row, 1, param).font = Font(bold=True)
            ws.cell(row, 2, value)
            ws.cell(row, 3, desc)

        # Add boolean validation
        bool_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        for row in [7, 8, 9]:  # Boolean fields
            bool_validation.add(f'B{row}')
        ws.add_data_validation(bool_validation)

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50

        # Apply borders
        for row in range(1, len(params) + 2):
            for col in range(1, 4):
                ws.cell(row, col).border = cls.BORDER

    @classmethod
    def _create_validation_sheet(cls, wb: Workbook, config: SegmentationConfig):
        """Create Validation Tests sheet."""
        ws = wb.create_sheet("Validation Tests")

        # Header
        ws['A1'] = "Test Name"
        ws['B1'] = "Include"
        ws['C1'] = "Description"

        for col in range(1, 4):
            ws.cell(1, col).font = cls.HEADER_FONT
            ws.cell(1, col).fill = cls.HEADER_FILL

        # Available tests
        available_tests = {
            'chi_squared': 'Chi-squared test for segment discrimination',
            'psi': 'Population Stability Index for temporal validation',
            'binomial': 'Binomial confidence intervals for PD estimates',
            'ks': 'Kolmogorov-Smirnov test',
            'gini': 'Gini coefficient for model discrimination'
        }

        row = 2
        for test, description in available_tests.items():
            ws.cell(row, 1, test)
            ws.cell(row, 2, "Yes" if test in config.irb_params.validation_tests else "No")
            ws.cell(row, 3, description)
            row += 1

        # Add boolean validation
        bool_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        bool_validation.add(f'B2:B{row-1}')
        ws.add_data_validation(bool_validation)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 60

        # Apply borders
        for r in range(1, row):
            for col in range(1, 4):
                ws.cell(r, col).border = cls.BORDER

    @classmethod
    def _create_help_sheet(cls, wb: Workbook):
        """Create Help sheet with parameter documentation."""
        ws = wb.create_sheet("Help")

        ws['A1'] = "IRB Segmentation - Parameter Guide"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        help_content = [
            ("", ""),
            ("Data Configuration", ""),
            ("Data Source", "Path to your CSV file, or use built-in loaders: german_credit, lending_club, taiwan_credit, home_credit"),
            ("Sample Size", "For testing, use a sample (e.g., 10000). Leave blank for full dataset in production."),
            ("", ""),
            ("IRB Parameters - Tree Structure", ""),
            ("Max Depth", "Tree depth. Smaller datasets: 2-3, Medium: 3-4, Large: 4-5. Higher depth = more segments but risk of overfitting."),
            ("Min Samples Leaf", "Minimum observations per segment. Should be 2-5% of dataset. Ensures segment stability."),
            ("Min Samples Split", "Must be at least 2 * Min Samples Leaf. Controls when nodes can split."),
            ("", ""),
            ("IRB Parameters - Basel Requirements", ""),
            ("Min Defaults Per Leaf", "Basel II/III recommends ≥20 defaults per segment for statistical reliability. Can be lower for very low default rate datasets."),
            ("Min Default Rate Diff", "Minimum PD difference between segments. 0.001 (0.1%) ensures meaningful discrimination."),
            ("Significance Level", "For statistical tests. 0.01 (1%) is standard. Lower = more stringent."),
            ("", ""),
            ("Density Controls", ""),
            ("Min Segment Density", "No segment should have <10% of observations (Basel guideline). Prevents tiny segments."),
            ("Max Segment Density", "No segment should have >50% of observations. Prevents one giant segment."),
            ("", ""),
            ("Business Constraints", ""),
            ("Forced Splits", "Mandatory thresholds (e.g., fico_score: 680 means always split at FICO 680). Enforces business rules or regulatory requirements."),
            ("Monotone Constraints", "Ensures risk ordering makes sense. Increasing = higher value = higher risk. Decreasing = higher value = lower risk."),
            ("", ""),
            ("Output", ""),
            ("Output Formats", "json (machine-readable), excel (for review), html (dashboard), yaml (version control)"),
            ("Create Dashboard", "Generate interactive HTML visualization of segments"),
            ("Create Excel Template", "Generate Excel file for threshold editing and segment merging"),
            ("", ""),
            ("Tips", ""),
            ("Start Conservative", "Use suggested parameters first, then adjust based on results."),
            ("Basel Compliance", "Green cells indicate Basel-compliant values. Yellow = caution. Red = non-compliant."),
            ("Validation Tests", "All tests should pass. If not, adjust parameters or review data quality."),
        ]

        row = 3
        for title, content in help_content:
            if not content:  # Empty line
                row += 1
                continue

            ws.cell(row, 1, title).font = Font(bold=True) if content else Font()
            ws.cell(row, 2, content)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 100


if __name__ == '__main__':
    # Test the export
    from irb_segmentation import create_default_config

    config = create_default_config(
        data_source='data/test.csv',
        data_type='csv'
    )
    config.name = "Test Configuration"
    config.description = "Testing Excel export"

    ConfigToExcel.export_config(config, 'test_config.xlsx', 'standard')
    print("Test complete!")
