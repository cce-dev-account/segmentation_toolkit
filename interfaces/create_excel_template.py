"""
Create Excel Template for User-Friendly Segment Modifications

Generates an Excel workbook with three worksheets:
1. Segment Actions - Merge/keep decisions
2. Threshold Editor - Business rule splits
3. Model Parameters - Parameter adjustments
"""

import json
import csv
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Note: openpyxl not available. Install with: pip install openpyxl")
    print("Falling back to CSV format.")


def simplify_path_conditions(rule_string):
    """
    Simplify redundant conditions in a decision path.

    Example: "int_rate <= 13.60 AND int_rate <= 10.58 AND int_rate <= 6.86"
    Becomes: [("int_rate", "<=", 6.86)]

    Also handles categorical conditions: "loan_purpose IN ['education', 'medical']"
    Becomes: [("loan_purpose", "IN", ['education', 'medical'])]

    Returns list of simplified conditions: [(feature, operator, value), ...]
    """
    conditions = rule_string.split(' AND ')

    # Group conditions by feature
    feature_conditions = {}
    categorical_conditions = []

    for condition in conditions:
        # Check for categorical condition (IN operator)
        if ' IN ' in condition:
            parts = condition.split(' IN ')
            if len(parts) == 2:
                feature = parts[0].strip()
                # Parse list of categories (e.g., "['A', 'B']")
                categories_str = parts[1].strip()
                try:
                    # Safely evaluate the list
                    import ast
                    categories = ast.literal_eval(categories_str)
                    categorical_conditions.append((feature, 'IN', categories))
                except:
                    # If parsing fails, keep as string
                    categorical_conditions.append((feature, 'IN', categories_str))
            continue

        # Parse numeric condition
        for op in ['<=', '>=', '<', '>', '==']:
            if op in condition:
                parts = condition.split(op)
                if len(parts) == 2:
                    feature = parts[0].strip()
                    try:
                        value = float(parts[1].strip())
                    except:
                        continue

                    if feature not in feature_conditions:
                        feature_conditions[feature] = {'<=': [], '>=': [], '<': [], '>': [], '==': []}

                    feature_conditions[feature][op].append(value)
                break

    # Simplify each feature's conditions
    simplified = []

    # Add categorical conditions first
    simplified.extend(categorical_conditions)

    for feature, ops in feature_conditions.items():
        # Collect all bounds
        upper_eq = ops['<=']  # upper bound inclusive
        lower_eq = ops['>=']  # lower bound inclusive
        upper = ops['<']      # upper bound exclusive
        lower = ops['>']      # lower bound exclusive
        equal = ops['==']

        if equal:
            # Exact value
            simplified.append((feature, '==', equal[0]))
        elif upper_eq or upper or lower_eq or lower:
            # Determine the most restrictive bounds
            min_upper_eq = min(upper_eq) if upper_eq else None
            min_upper = min(upper) if upper else None
            max_lower_eq = max(lower_eq) if lower_eq else None
            max_lower = max(lower) if lower else None

            # Determine actual bounds
            # Upper bound: take the most restrictive (minimum)
            if min_upper_eq is not None and min_upper is not None:
                if min_upper_eq <= min_upper:
                    upper_bound = (min_upper_eq, '<=')
                else:
                    upper_bound = (min_upper, '<')
            elif min_upper_eq is not None:
                upper_bound = (min_upper_eq, '<=')
            elif min_upper is not None:
                upper_bound = (min_upper, '<')
            else:
                upper_bound = None

            # Lower bound: take the most restrictive (maximum)
            if max_lower_eq is not None and max_lower is not None:
                if max_lower_eq >= max_lower:
                    lower_bound = (max_lower_eq, '>=')
                else:
                    lower_bound = (max_lower, '>')
            elif max_lower_eq is not None:
                lower_bound = (max_lower_eq, '>=')
            elif max_lower is not None:
                lower_bound = (max_lower, '>')
            else:
                lower_bound = None

            # Create simplified condition
            if lower_bound and upper_bound:
                # Range condition
                simplified.append((feature, 'range', (lower_bound[0], lower_bound[1], upper_bound[0], upper_bound[1])))
            elif upper_bound:
                # Only upper bound
                simplified.append((feature, upper_bound[1], upper_bound[0]))
            elif lower_bound:
                # Only lower bound
                simplified.append((feature, lower_bound[1], lower_bound[0]))

    return simplified


def create_csv_template(
    report_file: str = "lending_club_full_report.json",
    rules_file: str = "segment_rules_detailed.json",
    output_prefix: str = "modification_template"
):
    """
    Create CSV templates (Excel alternative that works everywhere).

    Creates 3 CSV files:
    1. segment_actions.csv
    2. threshold_editor.csv
    3. model_parameters.csv
    """

    print("\n" + "=" * 80)
    print("CREATING USER-FRIENDLY MODIFICATION TEMPLATES")
    print("=" * 80)

    # Load data
    report_path = Path(__file__).parent.parent / report_file
    rules_path = Path(__file__).parent.parent / rules_file

    if not report_path.exists():
        print(f"Error: {report_file} not found")
        return

    with open(report_path, 'r') as f:
        report = json.load(f)

    if rules_path.exists():
        with open(rules_path, 'r') as f:
            rules_data = json.load(f)
    else:
        print(f"Warning: {rules_file} not found. Creating basic template.")
        rules_data = {}

    stats = report['segment_statistics']
    params = report['parameters']
    feature_thresholds = rules_data.get('feature_thresholds', {})

    # Create output directory
    output_dir = Path(__file__).parent / 'templates'
    output_dir.mkdir(exist_ok=True)

    # 1. SEGMENT ACTIONS WORKSHEET
    print("\nCreating Segment Actions worksheet...")
    segment_actions_file = output_dir / f"{output_prefix}_segment_actions.csv"

    with open(segment_actions_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "SEGMENT MODIFICATION WORKSHEET",
            "",
            "",
            "",
            "",
            "",
            f"Generated from: {report_file}"
        ])
        writer.writerow([])
        writer.writerow([
            "Instructions:",
            "1. In 'Action' column, enter KEEP or MERGE",
            "2. If MERGE, specify target segment ID in 'Merge_Into' column",
            "3. Add notes explaining your decision",
            "4. Save this file and run: python interfaces/excel_to_json.py"
        ])
        writer.writerow([])

        # Column headers
        writer.writerow([
            "Segment_ID",
            "Observations",
            "Defaults",
            "PD_Rate_%",
            "Density_%",
            "Risk_Level",
            "Action",
            "Merge_Into",
            "Notes"
        ])

        # Data rows
        for seg_id in sorted([int(k) for k in stats.keys()]):
            s = stats[str(seg_id)]
            dr = s['default_rate'] * 100

            # Determine risk level
            if dr < 5:
                risk = "Very Low"
            elif dr < 10:
                risk = "Low"
            elif dr < 15:
                risk = "Medium"
            elif dr < 20:
                risk = "High"
            else:
                risk = "Very High"

            # Suggest merge for segments 1 and 4
            if seg_id == 1:
                action = "MERGE"
                merge_into = "4"
                notes = "Similar to Segment 4 (7.98% vs 8.35% PD)"
            elif seg_id == 4:
                action = "KEEP"
                merge_into = ""
                notes = "Accept merge from Segment 1"
            else:
                action = "KEEP"
                merge_into = ""
                notes = ""

            writer.writerow([
                seg_id,
                f"{s['n_observations']:,}",
                f"{s['n_defaults']:,}",
                f"{dr:.2f}",
                f"{s['density']*100:.2f}",
                risk,
                action,
                merge_into,
                notes
            ])

        writer.writerow([])
        writer.writerow([
            "Valid Actions:",
            "KEEP = Keep segment as-is",
            "MERGE = Merge this segment into another"
        ])

    print(f"  [OK] Created: {segment_actions_file}")

    # 2. THRESHOLD EDITOR WORKSHEET
    print("\nCreating Threshold Editor worksheet...")
    threshold_file = output_dir / f"{output_prefix}_thresholds.csv"

    with open(threshold_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "BUSINESS RULE THRESHOLDS",
            "",
            "",
            "",
            f"Generated from: {rules_file}"
        ])
        writer.writerow([])
        writer.writerow([
            "Instructions:",
            "1. Add rows for any forced splits you want",
            "2. Use exact feature names from 'Available_Features' section below",
            "3. Specify threshold value",
            "4. Provide business reason and regulatory flag"
        ])
        writer.writerow([])

        # Column headers
        writer.writerow([
            "Feature_Name",
            "Threshold_Value",
            "Business_Reason",
            "Regulatory_Requirement",
            "Priority"
        ])

        # Example rows
        writer.writerow([
            "int_rate",
            "15.0",
            "Company policy threshold for subprime",
            "No",
            "High"
        ])
        writer.writerow([
            "fico_range_high",
            "650",
            "Prime/subprime regulatory cutoff",
            "Yes",
            "High"
        ])
        writer.writerow([
            "annual_inc",
            "60000",
            "High-income segment for preferential pricing",
            "No",
            "Medium"
        ])
        writer.writerow([])

        # Current thresholds reference
        writer.writerow(["CURRENT THRESHOLDS (for reference):"])
        writer.writerow(["Feature", "Existing_Thresholds"])

        for feature, thresholds in sorted(feature_thresholds.items()):
            threshold_str = ", ".join([f"{t:.2f}" for t in thresholds])
            writer.writerow([feature, threshold_str])

        writer.writerow([])
        writer.writerow(["AVAILABLE FEATURES:"])
        writer.writerow(["Feature_Name", "Description"])

        feature_descriptions = {
            "int_rate": "Interest rate (%)",
            "fico_range_high": "FICO score (high range)",
            "fico_range_low": "FICO score (low range)",
            "annual_inc": "Annual income ($)",
            "dti": "Debt-to-income ratio (%)",
            "loan_amnt": "Loan amount ($)",
            "installment": "Monthly installment ($)",
            "inq_last_6mths": "Credit inquiries (last 6 months)"
        }

        for feature, desc in feature_descriptions.items():
            writer.writerow([feature, desc])

    print(f"  [OK] Created: {threshold_file}")

    # 3. MODEL PARAMETERS WORKSHEET
    print("\nCreating Model Parameters worksheet...")
    parameters_file = output_dir / f"{output_prefix}_parameters.csv"

    with open(parameters_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "MODEL PARAMETER ADJUSTMENTS",
            "",
            "",
            "",
            f"Generated from: {report_file}"
        ])
        writer.writerow([])
        writer.writerow([
            "Instructions:",
            "1. Modify 'New_Value' column with desired changes",
            "2. Leave 'New_Value' blank to keep current value",
            "3. Add reason for any changes"
        ])
        writer.writerow([])

        # Column headers
        writer.writerow([
            "Parameter",
            "Current_Value",
            "New_Value",
            "Valid_Range",
            "Description",
            "Change_Reason"
        ])

        # Parameter rows
        param_info = [
            ("max_depth", params.get('max_depth', 5), "", "3-7",
             "Maximum tree depth", ""),
            ("min_samples_leaf", params.get('min_samples_leaf', 10000), "", "1000-50000",
             "Minimum observations per segment", ""),
            ("min_defaults_per_leaf", params.get('min_defaults_per_leaf', 500), "", "50-2000",
             "Minimum defaults per segment", ""),
            ("min_segment_density", params.get('min_segment_density', 0.05), "", "0.01-0.20",
             "Minimum segment size (% of population)", ""),
            ("max_segment_density", params.get('max_segment_density', 0.40), "", "0.30-0.60",
             "Maximum segment size (% of population)", "")
        ]

        for param_name, current, new, valid_range, description, reason in param_info:
            writer.writerow([
                param_name,
                current,
                new,
                valid_range,
                description,
                reason
            ])

        writer.writerow([])
        writer.writerow(["PARAMETER GUIDANCE:"])
        writer.writerow(["Change", "Effect"])
        writer.writerow(["Increase max_depth", "More segments, more granular"])
        writer.writerow(["Decrease min_samples_leaf", "More segments, smaller sizes"])
        writer.writerow(["Increase min_defaults_per_leaf", "Fewer segments, higher statistical power"])
        writer.writerow(["Decrease min_segment_density", "Allow smaller segments"])
        writer.writerow(["Decrease max_segment_density", "Force more balanced segments"])

    print(f"  [OK] Created: {parameters_file}")

    # 4. README
    readme_file = output_dir / "README.txt"
    with open(readme_file, 'w') as f:
        f.write("""SEGMENTATION MODIFICATION TEMPLATES
====================================

You have 3 CSV files to edit:

1. modification_template_segment_actions.csv
   - Decide which segments to merge
   - Mark segments as KEEP or MERGE
   - Specify merge targets

2. modification_template_thresholds.csv
   - Add forced business rule splits
   - Specify feature thresholds (e.g., FICO 650)
   - Document regulatory requirements

3. modification_template_parameters.csv
   - Adjust model parameters
   - Change depth, minimums, density constraints
   - Document reasons for changes

WORKFLOW:
---------
1. Open files in Excel/LibreOffice/Google Sheets
2. Make your modifications
3. Save the files (keep CSV format)
4. Run: python interfaces/excel_to_json.py
5. This generates modification.json
6. Apply with: python apply_modifications.py modification.json

TIPS:
-----
- Edit one file at a time
- Start with segment_actions (merge similar segments)
- Then add thresholds (business rules)
- Finally adjust parameters (if needed)
- Always document your reasons in Notes column

HELP:
-----
- See MODELER_GUIDE.md for detailed instructions
- Check INTERFACE_COMPARISON.md for which interface to use
- View examples in interfaces/templates/examples/
""")

    print(f"  [OK] Created: {readme_file}")

    print("\n" + "=" * 80)
    print("TEMPLATES CREATED SUCCESSFULLY")
    print("=" * 80)
    print(f"\nLocation: {output_dir}")
    print("\nFiles created:")
    print(f"  1. {segment_actions_file.name}")
    print(f"  2. {threshold_file.name}")
    print(f"  3. {parameters_file.name}")
    print(f"  4. {readme_file.name}")
    print("\nNext steps:")
    print("  1. Open CSV files in Excel/spreadsheet software")
    print("  2. Make your modifications")
    print("  3. Save files")
    print("  4. Run: python interfaces/excel_to_json.py")

    return output_dir


def create_excel_workbook(
    report_file: str = "lending_club_full_report.json",
    rules_file: str = "segment_rules_detailed.json",
    output_file: str = "modification_template.xlsx"
):
    """Create Excel workbook with multiple worksheets."""

    if not EXCEL_AVAILABLE:
        print("\nExcel format not available. Using CSV instead.")
        return create_csv_template(report_file, rules_file)

    print("\n" + "=" * 80)
    print("CREATING EXCEL WORKBOOK WITH MODIFICATION TEMPLATES")
    print("=" * 80)

    # Load data
    report_path = Path(__file__).parent.parent / report_file
    rules_path = Path(__file__).parent.parent / rules_file

    if not report_path.exists():
        print(f"Error: {report_file} not found")
        return False

    with open(report_path, 'r') as f:
        report = json.load(f)

    if rules_path.exists():
        with open(rules_path, 'r') as f:
            rules_data = json.load(f)
    else:
        rules_data = {'feature_thresholds': {}}

    # Extract segment statistics
    segment_stats = {}
    pd_data = report.get('validation_results', {}).get('train', {}).get('validations', {}).get('binomial', {}).get('confidence_intervals', {})
    density_data = report.get('validation_results', {}).get('train', {}).get('validations', {}).get('density', {}).get('densities', {})

    for seg_id in sorted([k for k in pd_data.keys() if k.isdigit()], key=int):
        segment_stats[seg_id] = {
            'n_observations': pd_data[seg_id]['n_observations'],
            'n_defaults': pd_data[seg_id]['n_defaults'],
            'default_rate': pd_data[seg_id]['default_rate'],
            'density': density_data.get(seg_id, 0)
        }

    feature_thresholds = rules_data.get('feature_thresholds', {})
    params = report.get('parameters', {})

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    instruction_font = Font(italic=True, size=10)
    data_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    editable_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === WORKSHEET 1: Segment Actions ===
    ws1 = wb.create_sheet("Segment Actions")

    # Title
    ws1['A1'] = "SEGMENT MERGE/KEEP DECISIONS"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:I1')

    # Instructions
    ws1['A2'] = "Instructions:"
    ws1['A2'].font = Font(bold=True, size=11)
    ws1['A3'] = "1. Review segment statistics below"
    ws1['A4'] = "2. Set 'Action' column to MERGE or KEEP for each segment"
    ws1['A5'] = "3. For MERGE, specify target segment ID in 'Merge_Into' column"
    ws1['A6'] = "4. Add notes explaining your decision"
    for row in range(2, 7):
        ws1[f'A{row}'].fill = instruction_fill
        ws1.merge_cells(f'A{row}:I{row}')

    # Headers
    headers = ['Segment_ID', 'Observations', 'Defaults', 'PD_Rate_%', 'Density_%',
               'Risk_Level', 'Action', 'Merge_Into', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(8, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    row = 9
    for seg_id in sorted(segment_stats.keys(), key=int):
        stats = segment_stats[seg_id]
        pd_rate = stats['default_rate'] * 100
        density = stats['density'] * 100

        # Risk classification
        if pd_rate < 5:
            risk = "Very Low"
        elif pd_rate < 8:
            risk = "Low"
        elif pd_rate < 12:
            risk = "Low-Medium"
        elif pd_rate < 20:
            risk = "High"
        else:
            risk = "Very High"

        # Pre-fill recommendation: merge segment 1 into 4
        if seg_id == '1':
            action = "MERGE"
            merge_into = "4"
            notes = "Similar to Segment 4 (7.98% vs 8.35% PD)"
        else:
            action = "KEEP"
            merge_into = ""
            notes = ""

        ws1.cell(row, 1, int(seg_id))
        ws1.cell(row, 2, stats['n_observations'])
        ws1.cell(row, 3, stats['n_defaults'])
        ws1.cell(row, 4, round(pd_rate, 2))
        ws1.cell(row, 5, round(density, 2))
        ws1.cell(row, 6, risk)
        ws1.cell(row, 7, action)
        ws1.cell(row, 8, merge_into)
        ws1.cell(row, 9, notes)

        # Style data cells
        for col in range(1, 7):
            ws1.cell(row, col).fill = data_fill
            ws1.cell(row, col).border = border

        # Editable cells highlighted
        for col in range(7, 10):
            ws1.cell(row, col).fill = editable_fill
            ws1.cell(row, col).border = border

        row += 1

    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 50

    # === WORKSHEET 2: Segment Rules (Primary editing interface) ===
    ws_rules = wb.create_sheet("Segment Rules")

    # Title
    ws_rules['A1'] = "SEGMENT DECISION RULES - EDIT THRESHOLDS HERE"
    ws_rules['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_rules['A1'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws_rules.merge_cells('A1:F1')

    # Instructions
    ws_rules['A2'] = "Instructions:"
    ws_rules['A2'].font = Font(bold=True, size=11)
    ws_rules['A3'] = "• Each segment shows its complete decision paths below"
    ws_rules['A4'] = "• Edit threshold values in the 'New_Value' column to change segment boundaries"
    ws_rules['A5'] = "• Leave 'New_Value' blank to keep current threshold"
    ws_rules['A6'] = "• All conditions in a path are combined with AND logic"
    for r in range(2, 7):
        ws_rules[f'A{r}'].fill = instruction_fill
        ws_rules.merge_cells(f'A{r}:F{r}')

    current_row = 9

    # Get segment rules and statistics
    segment_rules = rules_data.get('segment_rules', {})

    for seg_id in sorted(segment_stats.keys(), key=int):
        stats = segment_stats[seg_id]
        seg_data = segment_rules.get(seg_id, {})
        paths = seg_data.get('rules', [])

        pd_rate = stats['default_rate'] * 100
        density = stats['density'] * 100

        # Risk classification
        if pd_rate < 5:
            risk = "Very Low"
            risk_color = "C6EFCE"
        elif pd_rate < 8:
            risk = "Low"
            risk_color = "E7F2F8"
        elif pd_rate < 12:
            risk = "Low-Medium"
            risk_color = "FFF4CE"
        elif pd_rate < 20:
            risk = "High"
            risk_color = "FFCCCC"
        else:
            risk = "Very High"
            risk_color = "FF9999"

        # Segment header
        ws_rules[f'A{current_row}'] = f"SEGMENT {seg_id}: {risk} Risk"
        ws_rules[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_rules[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_rules.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1

        # Segment statistics
        ws_rules[f'A{current_row}'] = f"PD: {pd_rate:.2f}% | Observations: {stats['n_observations']:,} | Defaults: {stats['n_defaults']:,} | Density: {density:.1f}%"
        ws_rules[f'A{current_row}'].fill = PatternFill(start_color=risk_color, end_color=risk_color, fill_type="solid")
        ws_rules.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1

        # Path headers
        ws_rules[f'A{current_row}'] = "Feature"
        ws_rules[f'B{current_row}'] = "Condition"
        ws_rules[f'C{current_row}'] = "Current Value"
        ws_rules[f'D{current_row}'] = "New Value"
        ws_rules[f'E{current_row}'] = "Format"
        ws_rules[f'F{current_row}'] = "Business Reason"

        for col in range(1, 7):
            ws_rules.cell(current_row, col).fill = header_fill
            ws_rules.cell(current_row, col).font = header_font
            ws_rules.cell(current_row, col).alignment = Alignment(horizontal='center', vertical='center')
            ws_rules.cell(current_row, col).border = border
        current_row += 1

        # Process each decision path
        for path_idx, rule in enumerate(paths, 1):
            # Simplify the rule to remove redundant conditions
            simplified_conditions = simplify_path_conditions(rule)

            # Path label
            if len(paths) > 1:
                ws_rules[f'A{current_row}'] = f"Path {path_idx}:"
                ws_rules[f'A{current_row}'].font = Font(bold=True, italic=True)
                ws_rules[f'A{current_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                ws_rules.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1

            for condition_tuple in simplified_conditions:
                feature = condition_tuple[0]
                op = condition_tuple[1]

                # Handle categorical conditions
                if op == 'IN':
                    categories = condition_tuple[2]
                    if isinstance(categories, list):
                        display_val = ', '.join([str(c) for c in categories])
                    else:
                        display_val = str(categories)

                    ws_rules[f'A{current_row}'] = feature
                    ws_rules[f'B{current_row}'] = "IN"
                    ws_rules[f'C{current_row}'] = display_val
                    ws_rules[f'D{current_row}'] = ""  # New value (user can edit)
                    ws_rules[f'E{current_row}'] = "categorical"
                    ws_rules[f'F{current_row}'] = ""  # Business reason

                    for col in range(1, 7):
                        ws_rules.cell(current_row, col).border = border
                    current_row += 1
                    continue

                # Determine format for numeric conditions
                if 'int_rate' in feature or 'dti' in feature:
                    fmt = '%'
                elif 'annual_inc' in feature or 'loan_amnt' in feature or 'installment' in feature:
                    fmt = '$'
                elif 'fico' in feature:
                    fmt = 'score'
                else:
                    fmt = 'number'

                # Handle range vs single bound
                if op == 'range':
                    # Range: (lower_val, lower_op, upper_val, upper_op)
                    range_data = condition_tuple[2]
                    lower_val, lower_op, upper_val, upper_op = range_data

                    # Format values
                    if fmt == '$':
                        lower_str = f"${lower_val:,.2f}"
                        upper_str = f"${upper_val:,.2f}"
                    elif fmt == '%':
                        lower_str = f"{lower_val:.2f}%"
                        upper_str = f"{upper_val:.2f}%"
                    elif fmt == 'score':
                        lower_str = f"{lower_val:.1f}"
                        upper_str = f"{upper_val:.1f}"
                    else:
                        lower_str = f"{lower_val:.2f}"
                        upper_str = f"{upper_val:.2f}"

                    display_val = f"{lower_str} to {upper_str}"
                    display_op = "between"

                else:
                    # Single bound
                    value = condition_tuple[2]

                    # Format value
                    if fmt == '$':
                        display_val = f"${value:,.2f}"
                    elif fmt == '%':
                        display_val = f"{value:.2f}%"
                    elif fmt == 'score':
                        display_val = f"{value:.1f}"
                    else:
                        display_val = f"{value:.2f}"

                    display_op = op

                # Add row
                ws_rules.cell(current_row, 1, feature).fill = data_fill
                ws_rules.cell(current_row, 2, display_op).fill = data_fill
                ws_rules.cell(current_row, 3, display_val).fill = data_fill
                ws_rules.cell(current_row, 4, '').fill = editable_fill  # New value
                ws_rules.cell(current_row, 5, fmt).fill = data_fill
                ws_rules.cell(current_row, 6, '').fill = editable_fill  # Reason

                for col in range(1, 7):
                    ws_rules.cell(current_row, col).border = border
                    ws_rules.cell(current_row, col).alignment = Alignment(horizontal='left', vertical='center')

                current_row += 1

        # Add spacing between segments
        current_row += 1

    # Column widths
    ws_rules.column_dimensions['A'].width = 20
    ws_rules.column_dimensions['B'].width = 10
    ws_rules.column_dimensions['C'].width = 18
    ws_rules.column_dimensions['D'].width = 18
    ws_rules.column_dimensions['E'].width = 10
    ws_rules.column_dimensions['F'].width = 40

    # === WORKSHEET 3: Threshold Overview ===
    ws_overview = wb.create_sheet("Threshold Overview")

    # Title
    ws_overview['A1'] = "THRESHOLD EDITOR - OVERVIEW"
    ws_overview['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_overview['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_overview.merge_cells('A1:F1')

    # Instructions with recommendation
    ws_overview['A2'] = "RECOMMENDED: Start with 'Segment Rules' worksheet (easier to visualize)"
    ws_overview['A2'].font = Font(bold=True, size=11, color="ED7D31")
    ws_overview['A2'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws_overview.merge_cells('A2:F2')

    ws_overview['A3'] = "Instructions:"
    ws_overview['A3'].font = Font(bold=True, size=11)
    ws_overview['A4'] = "• For most users: Use 'Segment Rules' tab to see complete segment paths"
    ws_overview['A5'] = "• For advanced users: Click feature names below for feature-level editing"
    ws_overview['A6'] = "• Each feature has its own worksheet for focused threshold management"
    for row in range(3, 7):
        ws_overview[f'A{row}'].fill = instruction_fill
        ws_overview.merge_cells(f'A{row}:F{row}')

    # Add link to Segment Rules
    ws_overview['A8'] = "→ Go to Segment Rules (Recommended)"
    ws_overview['A8'].font = Font(bold=True, color="ED7D31", size=12, underline="single")
    ws_overview['A8'].hyperlink = "#'Segment Rules'!A1"
    ws_overview['A8'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws_overview.merge_cells('A8:F8')

    # Summary table header
    ws_overview['A9'] = "Feature"
    ws_overview['B9'] = "# Thresholds"
    ws_overview['C9'] = "Range"
    ws_overview['D9'] = "Key Values"
    ws_overview['E9'] = "Format"
    ws_overview['F9'] = "Click to Edit →"

    for col in range(1, 7):
        ws_overview.cell(9, col).fill = header_fill
        ws_overview.cell(9, col).font = header_font
        ws_overview.cell(9, col).alignment = Alignment(horizontal='center', vertical='center')
        ws_overview.cell(9, col).border = border

    # Summary data
    row = 10
    feature_info = [
        ('int_rate', '%'),
        ('annual_inc', '$'),
        ('fico_range_high', 'score'),
        ('fico_range_low', 'score'),
        ('dti', '%'),
        ('inq_last_6mths', 'count'),
        ('installment', '$'),
        ('loan_amnt', '$')
    ]

    for feature, fmt in feature_info:
        if feature in feature_thresholds:
            thresholds = feature_thresholds[feature]
            if isinstance(thresholds, list):
                count = len(thresholds)
                min_val = min(thresholds)
                max_val = max(thresholds)
                if fmt == '$':
                    range_str = f"${min_val:,.0f} - ${max_val:,.0f}"
                    key_vals = f"${thresholds[0]:,.0f}, ${thresholds[-1]:,.0f}"
                elif fmt == '%':
                    range_str = f"{min_val:.2f}% - {max_val:.2f}%"
                    key_vals = f"{thresholds[0]:.2f}%, {thresholds[-1]:.2f}%"
                else:
                    range_str = f"{min_val:.1f} - {max_val:.1f}"
                    key_vals = f"{thresholds[0]:.1f}, {thresholds[-1]:.1f}"
            else:
                count = 1
                if fmt == '$':
                    range_str = f"${thresholds:,.0f}"
                    key_vals = f"${thresholds:,.0f}"
                elif fmt == '%':
                    range_str = f"{thresholds:.2f}%"
                    key_vals = f"{thresholds:.2f}%"
                else:
                    range_str = f"{thresholds:.1f}"
                    key_vals = f"{thresholds:.1f}"

            ws_overview.cell(row, 1, feature).fill = data_fill
            ws_overview.cell(row, 2, count).fill = data_fill
            ws_overview.cell(row, 3, range_str).fill = data_fill
            ws_overview.cell(row, 4, key_vals).fill = data_fill
            ws_overview.cell(row, 5, fmt).fill = data_fill
            ws_overview.cell(row, 6, f"→ {feature}").fill = editable_fill
            ws_overview.cell(row, 6).font = Font(color="0563C1", underline="single")

            for col in range(1, 7):
                ws_overview.cell(row, col).border = border

            # Add hyperlink to feature worksheet
            ws_overview.cell(row, 6).hyperlink = f"#{feature}!A1"

            row += 1

    # Column widths
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 14
    ws_overview.column_dimensions['C'].width = 22
    ws_overview.column_dimensions['D'].width = 22
    ws_overview.column_dimensions['E'].width = 12
    ws_overview.column_dimensions['F'].width = 18

    # === WORKSHEETS 3+: Individual Feature Thresholds ===
    # Create worksheet for each feature
    for feature, fmt in feature_info:
        if feature not in feature_thresholds:
            continue

        ws_feature = wb.create_sheet(feature)

        # Title with feature name
        ws_feature['A1'] = f"{feature.upper()} THRESHOLDS"
        ws_feature['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_feature['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_feature.merge_cells('A1:E1')

        # Instructions
        ws_feature['A2'] = "Instructions:"
        ws_feature['A2'].font = Font(bold=True, size=11)
        ws_feature['A3'] = "• Edit 'New_Value' to change thresholds (leave blank to keep current)"
        ws_feature['A4'] = "• Insert rows to add new thresholds"
        ws_feature['A5'] = "• Delete rows to remove thresholds"
        ws_feature['A6'] = "• Changed values override model-selected thresholds"
        for r in range(2, 7):
            ws_feature[f'A{r}'].fill = instruction_fill
            ws_feature.merge_cells(f'A{r}:E{r}')

        # Headers
        headers = ['Order', 'Current_Value', 'New_Value', 'Used_By_Segments', 'Business_Reason']
        for col, header in enumerate(headers, 1):
            cell = ws_feature.cell(8, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Get threshold values
        thresholds = feature_thresholds[feature]
        if not isinstance(thresholds, list):
            thresholds = [thresholds]

        # Determine which segments use each threshold (from segment rules)
        threshold_usage = {}
        for seg_id, seg_data in rules_data.get('segment_rules', {}).items():
            for rule in seg_data.get('rules', []):
                if feature in rule:
                    for threshold in thresholds:
                        threshold_str = f"{threshold:.2f}" if isinstance(threshold, float) else str(threshold)
                        if threshold_str in rule or f"{threshold:.1f}" in rule or f"{threshold:.0f}" in rule:
                            if threshold not in threshold_usage:
                                threshold_usage[threshold] = []
                            if seg_id not in threshold_usage[threshold]:
                                threshold_usage[threshold].append(seg_id)

        # Data rows
        row_num = 9
        for order, threshold in enumerate(sorted(thresholds), 1):
            # Format value based on feature type
            if fmt == '$':
                formatted_val = f"${threshold:,.2f}"
            elif fmt == '%':
                formatted_val = f"{threshold:.2f}%"
            else:
                formatted_val = f"{threshold:.2f}" if isinstance(threshold, float) else str(threshold)

            # Get segments using this threshold
            segs = threshold_usage.get(threshold, [])
            seg_str = ', '.join(sorted(segs)) if segs else ''

            ws_feature.cell(row_num, 1, order).fill = data_fill
            ws_feature.cell(row_num, 2, formatted_val).fill = data_fill
            ws_feature.cell(row_num, 3, '').fill = editable_fill  # New value - user edits
            ws_feature.cell(row_num, 4, seg_str).fill = data_fill
            ws_feature.cell(row_num, 5, '').fill = editable_fill  # Reason - user edits

            for col in range(1, 6):
                ws_feature.cell(row_num, col).border = border
                ws_feature.cell(row_num, col).alignment = Alignment(horizontal='center' if col <= 2 else 'left', vertical='center')

            row_num += 1

        # Add 3 blank rows for new thresholds
        for extra_row in range(3):
            ws_feature.cell(row_num, 1, '').fill = editable_fill
            ws_feature.cell(row_num, 2, '').fill = editable_fill
            ws_feature.cell(row_num, 3, '').fill = editable_fill
            ws_feature.cell(row_num, 4, '').fill = editable_fill
            ws_feature.cell(row_num, 5, '').fill = editable_fill
            for col in range(1, 6):
                ws_feature.cell(row_num, col).border = border
            row_num += 1

        # Column widths
        ws_feature.column_dimensions['A'].width = 8
        ws_feature.column_dimensions['B'].width = 18
        ws_feature.column_dimensions['C'].width = 18
        ws_feature.column_dimensions['D'].width = 18
        ws_feature.column_dimensions['E'].width = 45

        # Back to overview link
        ws_feature['A' + str(row_num + 2)] = "← Back to Overview"
        ws_feature['A' + str(row_num + 2)].font = Font(color="0563C1", underline="single", bold=True)
        ws_feature['A' + str(row_num + 2)].hyperlink = "#'Threshold Overview'!A1"
        ws_feature.merge_cells(f'A{row_num + 2}:E{row_num + 2}')

    # === WORKSHEET 3: Model Parameters ===
    ws3 = wb.create_sheet("Model Parameters")

    # Title
    ws3['A1'] = "MODEL PARAMETER ADJUSTMENTS"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:F1')

    # Instructions
    ws3['A2'] = "Instructions:"
    ws3['A2'].font = Font(bold=True, size=11)
    ws3['A3'] = "1. Modify 'New_Value' column with desired changes"
    ws3['A4'] = "2. Leave 'New_Value' blank to keep current value"
    ws3['A5'] = "3. Add reason for any changes in 'Change_Reason' column"
    for row in range(2, 6):
        ws3[f'A{row}'].fill = instruction_fill
        ws3.merge_cells(f'A{row}:F{row}')

    # Headers
    headers = ['Parameter', 'Current_Value', 'New_Value', 'Valid_Range', 'Description', 'Change_Reason']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(7, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Parameter data
    param_info = {
        'max_depth': ('3-7', 'Maximum tree depth'),
        'min_samples_leaf': ('1000-50000', 'Minimum observations per segment'),
        'min_defaults_per_leaf': ('50-2000', 'Minimum defaults per segment'),
        'min_segment_density': ('0.01-0.20', 'Minimum segment size (% of population)'),
        'max_segment_density': ('0.30-0.60', 'Maximum segment size (% of population)'),
    }

    row = 8
    for param, current_value in params.items():
        valid_range, description = param_info.get(param, ('', param))

        ws3.cell(row, 1, param).fill = data_fill
        ws3.cell(row, 2, current_value).fill = data_fill
        ws3.cell(row, 3, "").fill = editable_fill  # New value - user fills
        ws3.cell(row, 4, valid_range).fill = data_fill
        ws3.cell(row, 5, description).fill = data_fill
        ws3.cell(row, 6, "").fill = editable_fill  # Reason - user fills

        for col in range(1, 7):
            ws3.cell(row, col).border = border

        row += 1

    # Column widths
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 38
    ws3.column_dimensions['F'].width = 40

    # Save workbook
    output_path = Path(__file__).parent.parent / output_file
    wb.save(output_path)

    print(f"\n[OK] Created Excel workbook: {output_path.name}")
    print(f"  Size: {output_path.stat().st_size:,} bytes")
    print(f"  Worksheets: {len(wb.worksheets)} total")
    print(f"    - Segment Actions")
    print(f"    - Threshold Overview (with links to 8 feature sheets)")
    print(f"    - Individual sheets: int_rate, annual_inc, fico_range_high, etc.")
    print(f"    - Model Parameters")
    print("\n" + "=" * 80)
    print("NEXT STEPS")
    print("=" * 80)
    print(f"1. Open {output_file} in Excel")
    print("2. Navigate using tabs or hyperlinks in Threshold Overview")
    print("3. Edit values, add/remove rows in each feature sheet")
    print("4. Save the file")
    print("5. Convert to JSON: python interfaces/excel_to_json.py")
    print("6. Apply changes: python apply_modifications.py modification.json")

    return True


def main():
    """Main entry point."""
    # Try Excel format first, fall back to CSV if not available
    if EXCEL_AVAILABLE:
        create_excel_workbook()
    else:
        create_csv_template()


if __name__ == "__main__":
    main()
