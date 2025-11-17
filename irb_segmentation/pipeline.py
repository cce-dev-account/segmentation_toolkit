"""
Segmentation Pipeline Orchestrator

Unified orchestration layer that manages the entire segmentation workflow
from data loading through iterative threshold editing.
"""

import numpy as np
import pandas as pd
from typing import Dict, Optional, Tuple, Any
from pathlib import Path
import json
from datetime import datetime
import warnings

from .config import SegmentationConfig
from .engine import IRBSegmentationEngine
from .params import IRBSegmentationParams
from .logger import get_logger
from .validators import (
    validate_dataframe,
    validate_target_column,
    check_data_quality
)

# Module-level logger
logger = get_logger(__name__)


class SegmentationPipeline:
    """
    Orchestrates the complete segmentation workflow.

    Manages data loading, model fitting, template generation, modifications,
    and result comparison with data kept in memory throughout.

    Example:
        >>> config = SegmentationConfig.from_yaml('config.yaml')
        >>> pipeline = SegmentationPipeline(config)
        >>> pipeline.load_data()
        >>> pipeline.fit_baseline()
        >>> template_path = pipeline.export_template()
        >>> # User edits template
        >>> pipeline.apply_modifications(template_path)
        >>> pipeline.compare_results()
        >>> pipeline.export_all()
    """

    def __init__(self, config: SegmentationConfig):
        """
        Initialize pipeline with configuration.

        Args:
            config: SegmentationConfig object
        """
        self.config = config
        self.verbose = config.verbose

        # Data storage (loaded once, reused)
        self.X_train = None
        self.y_train = None
        self.X_val = None
        self.y_val = None
        self.X_oot = None
        self.y_oot = None
        self.feature_names = None
        self.X_categorical = None

        # Model storage
        self.baseline_engine = None
        self.modified_engine = None

        # State tracking
        self.state = {
            'data_loaded': False,
            'baseline_fitted': False,
            'template_exported': False,
            'modifications_applied': False,
            'results_compared': False
        }

        # Results storage
        self.baseline_results = {}
        self.modified_results = {}
        self.comparison_results = {}

        # Ensure output directory exists
        Path(self.config.output.output_dir).mkdir(parents=True, exist_ok=True)

        if self.verbose:
            logger.info("=" * 70)
            logger.info("SEGMENTATION PIPELINE INITIALIZED")
            logger.info("=" * 70)
            logger.info(self.config.summary())

    def load_data(self) -> None:
        """
        Load data from configured source and keep in memory.

        Data is loaded once and reused across all pipeline stages.
        """
        if self.state['data_loaded']:
            if self.verbose:
                logger.info("\n[INFO] Data already loaded, skipping")
            return

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 1: LOADING DATA")
            logger.info("=" * 70)

        data_config = self.config.data

        # Load based on data type
        if data_config.data_type == 'csv':
            self._load_from_csv()
        elif data_config.data_type in ['german_credit', 'lending_club', 'taiwan_credit', 'home_credit']:
            self._load_from_loader()
        else:
            raise ValueError(f"Unsupported data_type: {data_config.data_type}")

        self.state['data_loaded'] = True

        if self.verbose:
            logger.info(f"\n[OK] Data loaded successfully")
            logger.info(f"     Training: {len(self.X_train):,} observations, {self.y_train.sum():,} defaults ({self.y_train.mean():.4f})")
            logger.info(f"     Validation: {len(self.X_val):,} observations, {self.y_val.sum():,} defaults ({self.y_val.mean():.4f})")
            if self.X_oot is not None:
                logger.info(f"     OOT: {len(self.X_oot):,} observations")
            logger.info(f"     Features: {len(self.feature_names)}")

    def _load_from_csv(self) -> None:
        """Load data from CSV file."""
        from sklearn.model_selection import train_test_split

        if self.verbose:
            logger.info(f"\nLoading CSV: {self.config.data.source}")

        df = pd.read_csv(self.config.data.source)

        if self.verbose:
            logger.info(f"Loaded {len(df):,} rows, {len(df.columns)} columns")

        # Validate DataFrame
        logger.info("\nValidating DataFrame...")
        df_report = validate_dataframe(
            df,
            min_rows=100,
            max_missing_pct=0.50
        )

        if self.verbose:
            # Show key data quality metrics
            if df_report.warnings:
                logger.info(f"Data quality warnings: {len(df_report.warnings)}")
                for warning in df_report.warnings[:3]:  # Show first 3
                    logger.info(f"  - {warning}")

        # Apply sampling if configured
        if self.config.data.sample_size:
            if len(df) > self.config.data.sample_size:
                df = df.sample(n=self.config.data.sample_size, random_state=self.config.data.random_state)
                if self.verbose:
                    logger.info(f"Sampled down to {len(df):,} rows")

        # Identify target column
        target_col = None

        # First, check if user explicitly specified target column
        if self.config.data.target_column:
            if self.config.data.target_column in df.columns:
                target_col = self.config.data.target_column
                if self.verbose:
                    logger.info(f"Using configured target column: {target_col}")
            else:
                raise ValueError(f"Configured target_column '{self.config.data.target_column}' not found in CSV. "
                               f"Available columns: {list(df.columns)}")

        # Otherwise, try common target column names
        if target_col is None:
            for col_name in ['default', 'target', 'y', 'label', 'loan_status']:
                if col_name in df.columns:
                    target_col = col_name
                    if self.verbose:
                        logger.info(f"Auto-detected target column: {target_col}")
                    break

        if target_col is None:
            # Use last column as target
            target_col = df.columns[-1]
            warnings.warn(f"No standard target column found, using last column: {target_col}")

        # Separate features and target
        y = df[target_col].values
        X_df = df.drop(columns=[target_col])

        # Handle missing values in target
        nan_mask = pd.isna(y)
        if nan_mask.any():
            n_missing = nan_mask.sum()
            if self.verbose:
                logger.info(f"Warning: Dropping {n_missing:,} rows with missing target values ({n_missing/len(y)*100:.2f}%)")
            # Filter out rows with missing target
            valid_mask = ~nan_mask
            y = y[valid_mask]
            X_df = X_df[valid_mask].reset_index(drop=True)

        # Validate and clean target variable
        unique_values = np.unique(y)
        if self.verbose:
            logger.info(f"\nTarget variable '{target_col}':")
            logger.info(f"  Unique values: {unique_values}")
            for val in unique_values:
                count = np.sum(y == val)
                logger.info(f"  Class {val}: {count:,} samples ({count/len(y)*100:.2f}%)")

        # Convert to binary if needed
        if len(unique_values) == 2:
            # Map to 0/1 if not already
            if not (set(unique_values) == {0, 1} or set(unique_values) == {0.0, 1.0}):
                if self.verbose:
                    logger.info(f"  Converting to binary: {unique_values[0]} -> 0, {unique_values[1]} -> 1")
                y = (y == unique_values[1]).astype(int)
        elif len(unique_values) > 2:
            # Multi-class - convert to binary by treating smallest class as positive
            warnings.warn(f"Target has {len(unique_values)} classes. Converting to binary (smallest class = 1)")
            class_counts = [(val, np.sum(y == val)) for val in unique_values]
            class_counts.sort(key=lambda x: x[1])
            positive_class = class_counts[0][0]
            y = (y == positive_class).astype(int)
            if self.verbose:
                logger.info(f"  Treating class '{positive_class}' as positive (1), others as negative (0)")

        # Remove classes with too few samples (< 2)
        unique_after, counts = np.unique(y, return_counts=True)
        if np.any(counts < 2):
            rare_classes = unique_after[counts < 2]
            if self.verbose:
                logger.info(f"Warning: Removing {np.sum(counts < 2)} samples from rare classes: {rare_classes}")
            valid_mask = np.isin(y, unique_after[counts >= 2])
            y = y[valid_mask]
            X_df = X_df.iloc[valid_mask].reset_index(drop=True)

        # Extract categorical features if specified
        if self.config.data.categorical_columns:
            categorical_cols = [col for col in self.config.data.categorical_columns if col in X_df.columns]
            if categorical_cols:
                if self.verbose:
                    logger.info(f"Extracting categorical features: {categorical_cols}")
                # Store categorical features before they're encoded
                self.X_categorical = {col: X_df[col].values for col in categorical_cols}
                # Remove from numeric features (will be encoded separately if needed)
                X_df = X_df.drop(columns=categorical_cols)
            else:
                self.X_categorical = None
                if self.verbose:
                    logger.info(f"Warning: No categorical columns found in data from config: {self.config.data.categorical_columns}")
        else:
            self.X_categorical = None

        # Select only numeric columns
        numeric_cols = X_df.select_dtypes(include=[np.number]).columns.tolist()
        X = X_df[numeric_cols].fillna(0).values
        self.feature_names = numeric_cols

        # Train/validation split with conditional stratification
        # Check if stratification is possible (all classes have >= 2 samples)
        unique_classes, class_counts = np.unique(y, return_counts=True)
        can_stratify = np.all(class_counts >= 2)

        if can_stratify:
            if self.verbose:
                logger.info(f"\nUsing stratified train/test split")
            self.X_train, self.X_val, self.y_train, self.y_val = train_test_split(
                X, y,
                test_size=0.3,
                random_state=self.config.data.random_state,
                stratify=y
            )
        else:
            if self.verbose:
                logger.info(f"\nWarning: Cannot stratify - some classes have < 2 samples. Using random split.")
            self.X_train, self.X_val, self.y_train, self.y_val = train_test_split(
                X, y,
                test_size=0.3,
                random_state=self.config.data.random_state
            )

        # No OOT for CSV (unless temporal split is implemented)
        self.X_oot = None
        self.y_oot = None

    def _load_from_loader(self) -> None:
        """Load data using built-in data loaders."""
        loader_name = self.config.data.data_type

        if self.verbose:
            logger.info(f"\nLoading data using loader: {loader_name}")

        if loader_name == 'german_credit':
            from data_loaders import load_german_credit
            result = load_german_credit()
        elif loader_name == 'lending_club':
            from data_loaders import load_lending_club
            sample_size = self.config.data.sample_size
            result = load_lending_club(
                sample_size=sample_size,
                use_oot=self.config.data.use_oot
            )
        elif loader_name == 'taiwan_credit':
            from data_loaders import load_taiwan_credit
            result = load_taiwan_credit()
        elif loader_name == 'home_credit':
            from data_loaders import load_home_credit
            result = load_home_credit()
        else:
            raise ValueError(f"Unknown loader: {loader_name}")

        # Unpack results (8-tuple now includes X_categorical)
        self.X_train, self.y_train, self.X_val, self.y_val, self.X_oot, self.y_oot, self.feature_names, self.X_categorical = result

    def fit_baseline(self) -> None:
        """
        Fit baseline segmentation model using configured parameters.
        """
        if not self.state['data_loaded']:
            raise RuntimeError("Must call load_data() before fit_baseline()")

        if self.state['baseline_fitted']:
            if self.verbose:
                logger.info("\n[INFO] Baseline already fitted, skipping")
            return

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 2: FITTING BASELINE MODEL")
            logger.info("=" * 70)

        # Create engine with configured parameters
        self.baseline_engine = IRBSegmentationEngine(self.config.irb_params)

        # Fit model
        self.baseline_engine.fit(
            X=self.X_train,
            y=self.y_train,
            X_val=self.X_val,
            y_val=self.y_val,
            feature_names=self.feature_names,
            X_categorical=self.X_categorical
        )

        # Store results
        self.baseline_results = self.baseline_engine.get_validation_report()

        self.state['baseline_fitted'] = True

        if self.verbose:
            logger.info(f"\n[OK] Baseline model fitted successfully")
            logger.info(f"     Segments: {len(np.unique(self.baseline_engine.segments_train_))}")
            logger.info(f"     Validation: {'PASSED' if self.baseline_results['validation_results']['train']['all_passed'] else 'FAILED'}")

    def export_template(self, format: str = 'excel') -> str:
        """
        Export template for threshold editing.

        Args:
            format: Template format ('excel' or 'json')

        Returns:
            Path to exported template file
        """
        if not self.state['baseline_fitted']:
            raise RuntimeError("Must call fit_baseline() before export_template()")

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 3: EXPORTING TEMPLATE")
            logger.info("=" * 70)

        output_dir = Path(self.config.output.output_dir)

        # For now, only JSON templates are supported
        # Excel template requires additional development
        if format == 'excel':
            if self.verbose:
                logger.info("  Excel template format requested, using JSON instead")
            format = 'json'

        if format == 'json':
            # Export JSON template
            template_name = self.config.output.template_name or "modification_template.json"
            template_path = output_dir / template_name

            # Create modification template
            template = {
                'metadata': {
                    'created': datetime.now().isoformat(),
                    'segments': len(np.unique(self.baseline_engine.segments_train_)),
                    'dataset': self.config.data.source
                },
                'modifications': {
                    'merge_segments': {
                        'description': 'List segment pairs to merge: [[seg1, seg2], [seg3, seg4]]',
                        'value': []
                    },
                    'forced_splits': {
                        'description': 'Add forced split points: {feature_name: threshold}',
                        'value': dict(self.config.irb_params.forced_splits)
                    },
                    'parameter_changes': {
                        'description': 'Modify parameters (triggers re-training)',
                        'max_depth': self.config.irb_params.max_depth,
                        'min_samples_leaf': self.config.irb_params.min_samples_leaf,
                        'min_defaults_per_leaf': self.config.irb_params.min_defaults_per_leaf,
                        'min_segment_density': self.config.irb_params.min_segment_density,
                        'max_segment_density': self.config.irb_params.max_segment_density
                    }
                },
                'current_segments': self.baseline_engine._get_segment_statistics()
            }

            with open(template_path, 'w') as f:
                json.dump(template, f, indent=2)

        self.state['template_exported'] = True

        if self.verbose:
            logger.info(f"\n[OK] Template exported: {template_path}")
            logger.info(f"\n     Edit the template and then run:")
            logger.info(f"     pipeline.apply_modifications('{template_path}')")

        return str(template_path)

    def apply_modifications(self, template_path: str) -> None:
        """
        Apply modifications from edited template and refit model.

        Args:
            template_path: Path to edited template file
        """
        if not self.state['template_exported']:
            warnings.warn("Template was not exported by this pipeline, proceeding anyway")

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 4: APPLYING MODIFICATIONS")
            logger.info("=" * 70)
            logger.info(f"\nReading modifications from: {template_path}")

        template_path = Path(template_path)

        if template_path.suffix == '.json':
            self._apply_json_modifications(template_path)
        elif template_path.suffix == '.xlsx':
            self._apply_excel_modifications(template_path)
        else:
            raise ValueError(f"Unsupported template format: {template_path.suffix}")

        self.state['modifications_applied'] = True

        if self.verbose:
            logger.info(f"\n[OK] Modifications applied successfully")
            logger.info(f"     Segments: {len(np.unique(self.modified_engine.segments_train_))}")

    def _apply_json_modifications(self, template_path: Path) -> None:
        """Apply modifications from JSON template."""
        with open(template_path, 'r') as f:
            template = json.load(f)

        mods = template['modifications']

        # Create modified parameters
        modified_params = IRBSegmentationParams(
            max_depth=mods['parameter_changes'].get('max_depth', self.config.irb_params.max_depth),
            min_samples_split=mods['parameter_changes'].get('min_samples_leaf', self.config.irb_params.min_samples_leaf) * 2,
            min_samples_leaf=mods['parameter_changes'].get('min_samples_leaf', self.config.irb_params.min_samples_leaf),
            min_defaults_per_leaf=mods['parameter_changes'].get('min_defaults_per_leaf', self.config.irb_params.min_defaults_per_leaf),
            min_segment_density=mods['parameter_changes'].get('min_segment_density', self.config.irb_params.min_segment_density),
            max_segment_density=mods['parameter_changes'].get('max_segment_density', self.config.irb_params.max_segment_density),
            forced_splits=mods['forced_splits']['value']
        )

        # Fit new model
        self.modified_engine = IRBSegmentationEngine(modified_params)
        self.modified_engine.fit(
            X=self.X_train,
            y=self.y_train,
            X_val=self.X_val,
            y_val=self.y_val,
            feature_names=self.feature_names,
            X_categorical=self.X_categorical
        )

        # Apply manual merges if specified
        merge_pairs = mods['merge_segments']['value']
        if merge_pairs:
            if self.verbose:
                logger.info(f"\nApplying {len(merge_pairs)} manual merges...")

            segments = self.modified_engine.segments_train_.copy()

            for seg1, seg2 in merge_pairs:
                if seg1 in segments and seg2 in segments:
                    segments[segments == seg2] = seg1

            # Re-label contiguously
            unique_segs = sorted(np.unique(segments))
            segment_map = {old: new for new, old in enumerate(unique_segs)}
            segments = np.array([segment_map[s] for s in segments])

            self.modified_engine.segments_train_ = segments

            # Re-validate
            from .validators import SegmentValidator
            self.modified_engine.validation_results_['train'] = SegmentValidator.run_all_validations(
                segments, self.y_train, modified_params
            )

        # Store results
        self.modified_results = self.modified_engine.get_validation_report()

    def _apply_excel_modifications(self, template_path: Path) -> None:
        """Apply modifications from Excel template."""
        # Import Excel converter
        import sys
        from pathlib import Path as P
        sys.path.insert(0, str(P(__file__).parent.parent / "interfaces"))

        try:
            from interfaces.excel_to_json import excel_to_json
            # Convert Excel to JSON
            json_path = template_path.with_suffix('.json')
            excel_to_json(str(template_path), str(json_path))
            # Apply JSON modifications
            self._apply_json_modifications(json_path)
        except ImportError:
            raise RuntimeError("Excel modification support requires interfaces.excel_to_json module")

    def compare_results(self) -> Dict:
        """
        Compare baseline and modified results.

        Returns:
            Dictionary with comparison metrics
        """
        if not self.state['baseline_fitted']:
            raise RuntimeError("Baseline model not fitted")

        if not self.state['modifications_applied']:
            warnings.warn("No modifications applied, comparing baseline to itself")
            self.modified_engine = self.baseline_engine
            self.modified_results = self.baseline_results

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 5: COMPARING RESULTS")
            logger.info("=" * 70)

        baseline_stats = self.baseline_results['segment_statistics']
        modified_stats = self.modified_results['segment_statistics']

        comparison = {
            'baseline': {
                'n_segments': len(baseline_stats),
                'segments': baseline_stats
            },
            'modified': {
                'n_segments': len(modified_stats),
                'segments': modified_stats
            },
            'changes': {
                'segment_count_change': len(modified_stats) - len(baseline_stats)
            }
        }

        self.comparison_results = comparison
        self.state['results_compared'] = True

        if self.verbose:
            logger.info(f"\nSegments: {len(baseline_stats)} → {len(modified_stats)} ({len(modified_stats) - len(baseline_stats):+d})")
            logger.info(f"\nBaseline validation: {'PASSED' if self.baseline_results['validation_results']['train']['all_passed'] else 'FAILED'}")
            logger.info(f"Modified validation: {'PASSED' if self.modified_results['validation_results']['train']['all_passed'] else 'FAILED'}")

        return comparison

    def export_all(self) -> Dict[str, str]:
        """
        Export all outputs in configured formats.

        Returns:
            Dictionary mapping output type to file paths
        """
        if not self.state['baseline_fitted']:
            raise RuntimeError("Must fit baseline before exporting")

        if self.verbose:
            logger.info("=" * 70)
            logger.info("STAGE 6: EXPORTING OUTPUTS")
            logger.info("=" * 70)

        output_dir = Path(self.config.output.output_dir)
        exported_files = {}

        # Export baseline report
        if 'json' in self.config.output.output_formats:
            baseline_report = output_dir / "baseline_report.json"
            self.baseline_engine.export_report(str(baseline_report))
            exported_files['baseline_report'] = str(baseline_report)

        # Export modified report if available
        if self.state['modifications_applied'] and 'json' in self.config.output.output_formats:
            modified_report = output_dir / "modified_report.json"
            self.modified_engine.export_report(str(modified_report))
            exported_files['modified_report'] = str(modified_report)

        # Export comparison
        if self.state['results_compared'] and 'json' in self.config.output.output_formats:
            comparison_path = output_dir / "comparison_report.json"
            with open(comparison_path, 'w') as f:
                json.dump(self.comparison_results, f, indent=2)
            exported_files['comparison'] = str(comparison_path)

        # Export tree structure
        if 'json' in self.config.output.output_formats:
            tree_path = output_dir / "tree_structure.json"
            self.baseline_engine.export_tree_to_file(str(tree_path))
            exported_files['tree_structure'] = str(tree_path)

        # Export segment rules (human-readable text)
        if self.config.output.extract_rules:
            rules_path = output_dir / "segment_rules.txt"
            try:
                if self.verbose:
                    logger.info(f"Extracting segment rules: {rules_path}")

                # Get segment statistics from baseline engine
                segment_ids = np.unique(self.baseline_engine.segments_train_)

                with open(rules_path, 'w') as f:
                    f.write("=" * 80 + "\n")
                    f.write("SEGMENT DEFINITIONS AND STATISTICS\n")
                    f.write("=" * 80 + "\n\n")

                    # Get rules
                    rules = self.baseline_engine.get_segment_rules()

                    # Calculate statistics for each segment
                    for seg_id in sorted(segment_ids):
                        seg_mask_train = self.baseline_engine.segments_train_ == seg_id
                        n_obs = np.sum(seg_mask_train)
                        n_defaults = np.sum(self.y_train[seg_mask_train])
                        pd_rate = n_defaults / n_obs if n_obs > 0 else 0
                        density = n_obs / len(self.y_train) * 100

                        f.write(f"\nSegment {seg_id}:\n")
                        f.write(f"  Observations: {n_obs:,}\n")
                        f.write(f"  Defaults: {n_defaults:,} ({pd_rate:.2%})\n")
                        f.write(f"  Density: {density:.2f}%\n")

                        # Find matching rule
                        matching_rule = [r for r in rules if f"Segment {seg_id}" in r]
                        if matching_rule:
                            rule_condition = matching_rule[0].replace(f" THEN Segment {seg_id}", "")
                            f.write(f"  Rule: {rule_condition}\n")
                        f.write("\n" + "-" * 80 + "\n")

                    # Summary statistics
                    f.write("\n" + "=" * 80 + "\n")
                    f.write("OVERALL STATISTICS\n")
                    f.write("=" * 80 + "\n\n")
                    f.write(f"Total Observations (Training): {len(self.y_train):,}\n")
                    f.write(f"Total Defaults: {np.sum(self.y_train):,}\n")
                    f.write(f"Overall Default Rate: {np.mean(self.y_train):.2%}\n")
                    f.write(f"Number of Segments: {len(segment_ids)}\n")

                exported_files['segment_rules'] = str(rules_path)

                if self.verbose:
                    logger.info(f"  Segment rules exported: {rules_path}")
            except Exception as e:
                warnings.warn(f"Could not export segment rules: {e}")

        # Export comprehensive segment report with visualizations
        if self.config.output.extract_rules:
            try:
                if self.verbose:
                    logger.info(f"\nGenerating comprehensive segment report with visualizations...")

                # Import visualization module
                import sys
                sys.path.insert(0, str(Path(__file__).parent.parent))
                from scripts.visualize_segments import SegmentVisualizer

                # Create visualizer
                visualizer = SegmentVisualizer(
                    X=self.X_train,
                    y=self.y_train,
                    segments=self.baseline_engine.segments_train_,
                    feature_names=self.feature_names,
                    output_dir=str(output_dir)
                )

                # Generate all visualizations
                viz_outputs = visualizer.generate_all_visualizations(prefix="segment_viz")

                # Get comprehensive descriptions
                segment_descriptions = self.baseline_engine.get_comprehensive_segment_descriptions()

                # Generate markdown report
                markdown_path = output_dir / "segment_report.md"
                self._generate_markdown_report(markdown_path, segment_descriptions, viz_outputs)

                exported_files['segment_report_md'] = str(markdown_path)
                exported_files.update({f'viz_{k}': v for k, v in viz_outputs.items()})

                if self.verbose:
                    logger.info(f"  Comprehensive segment report created: {markdown_path}")
                    logger.info(f"  Generated {len(viz_outputs)} visualization files")

            except Exception as e:
                warnings.warn(f"Could not generate comprehensive segment report: {e}")
                import traceback
                if self.verbose:
                    traceback.print_exc()

        # Export Excel report
        if 'excel' in self.config.output.output_formats:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                excel_path = output_dir / "segmentation_report.xlsx"

                if self.verbose:
                    logger.info(f"Generating Excel report: {excel_path}")

                # Create workbook with segment summary
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Segment Summary"

                # Header
                ws['A1'] = "IRB Segmentation Analysis - Segment Summary"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')

                # Column headers
                headers = ['Segment ID', 'Observations', 'Defaults', 'Default Rate', 'Density (%)', 'Risk Level']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(3, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

                # Data rows
                segment_ids = np.unique(self.baseline_engine.segments_train_)
                for row_idx, seg_id in enumerate(sorted(segment_ids), 4):
                    seg_mask = self.baseline_engine.segments_train_ == seg_id
                    n_obs = np.sum(seg_mask)
                    n_defaults = np.sum(self.y_train[seg_mask])
                    pd_rate = n_defaults / n_obs if n_obs > 0 else 0
                    density = n_obs / len(self.y_train) * 100

                    # Determine risk level
                    if pd_rate < 0.10:
                        risk_level = "Low"
                    elif pd_rate < 0.20:
                        risk_level = "Medium"
                    else:
                        risk_level = "High"

                    ws.cell(row_idx, 1, int(seg_id))
                    ws.cell(row_idx, 2, int(n_obs))
                    ws.cell(row_idx, 3, int(n_defaults))
                    ws.cell(row_idx, 4, f"{pd_rate:.2%}")
                    ws.cell(row_idx, 5, f"{density:.2f}%")
                    ws.cell(row_idx, 6, risk_level)

                # Summary row
                summary_row = row_idx + 2
                ws.cell(summary_row, 1, "TOTAL").font = Font(bold=True)
                ws.cell(summary_row, 2, int(len(self.y_train)))
                ws.cell(summary_row, 3, int(np.sum(self.y_train)))
                ws.cell(summary_row, 4, f"{np.mean(self.y_train):.2%}")
                ws.cell(summary_row, 5, "100.00%")

                # Auto-size columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column].width = max_length + 2

                # Save
                wb.save(excel_path)
                exported_files['excel_report'] = str(excel_path)

                if self.verbose:
                    logger.info(f"  Excel report created: {excel_path}")
            except ImportError as e:
                warnings.warn(f"Could not create Excel report: openpyxl not available")
            except Exception as e:
                warnings.warn(f"Error creating Excel report: {e}")

        # Export HTML dashboard
        if 'html' in self.config.output.output_formats:
            try:
                html_path = output_dir / "dashboard.html"

                if self.verbose:
                    logger.info(f"Generating HTML dashboard: {html_path}")

                # Prepare segment data
                segment_ids = np.unique(self.baseline_engine.segments_train_)
                segment_data = []
                for seg_id in sorted(segment_ids):
                    seg_mask = self.baseline_engine.segments_train_ == seg_id
                    n_obs = int(np.sum(seg_mask))
                    n_defaults = int(np.sum(self.y_train[seg_mask]))
                    pd_rate = float(n_defaults / n_obs if n_obs > 0 else 0)
                    density = float(n_obs / len(self.y_train) * 100)

                    segment_data.append({
                        'id': int(seg_id),
                        'observations': n_obs,
                        'defaults': n_defaults,
                        'pd_rate': pd_rate,
                        'density': density
                    })

                # Create HTML
                html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>IRB Segmentation Dashboard</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 3px solid #366092; padding-bottom: 10px; }}
        h2 {{ color: #366092; margin-top: 30px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #366092; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background: #f5f5f5; }}
        .stat-box {{ display: inline-block; margin: 10px 20px; padding: 20px; background: #f9f9f9; border-left: 4px solid #366092; }}
        .stat-value {{ font-size: 24px; font-weight: bold; color: #366092; }}
        .stat-label {{ color: #666; font-size: 14px; }}
        .high-risk {{ color: #d32f2f; font-weight: bold; }}
        .medium-risk {{ color: #f57c00; font-weight: bold; }}
        .low-risk {{ color: #388e3c; font-weight: bold; }}
        .chart-container {{ margin: 30px 0; }}
        .bar {{ background: #366092; height: 30px; margin: 5px 0; color: white; padding: 5px 10px; border-radius: 3px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>IRB Segmentation Analysis Dashboard</h1>

        <div class="stat-box">
            <div class="stat-value">{len(segment_ids)}</div>
            <div class="stat-label">Segments</div>
        </div>
        <div class="stat-box">
            <div class="stat-value">{len(self.y_train):,}</div>
            <div class="stat-label">Total Observations</div>
        </div>
        <div class="stat-box">
            <div class="stat-value">{int(np.sum(self.y_train)):,}</div>
            <div class="stat-label">Total Defaults</div>
        </div>
        <div class="stat-box">
            <div class="stat-value">{np.mean(self.y_train):.2%}</div>
            <div class="stat-label">Overall Default Rate</div>
        </div>

        <h2>Segment Summary</h2>
        <table>
            <thead>
                <tr>
                    <th>Segment</th>
                    <th>Observations</th>
                    <th>Defaults</th>
                    <th>Default Rate</th>
                    <th>Density</th>
                    <th>Risk Level</th>
                </tr>
            </thead>
            <tbody>
"""

                for seg in segment_data:
                    if seg['pd_rate'] < 0.10:
                        risk_class = "low-risk"
                        risk_level = "Low"
                    elif seg['pd_rate'] < 0.20:
                        risk_class = "medium-risk"
                        risk_level = "Medium"
                    else:
                        risk_class = "high-risk"
                        risk_level = "High"

                    html_content += f"""
                <tr>
                    <td>{seg['id']}</td>
                    <td>{seg['observations']:,}</td>
                    <td>{seg['defaults']:,}</td>
                    <td>{seg['pd_rate']:.2%}</td>
                    <td>{seg['density']:.2f}%</td>
                    <td class="{risk_class}">{risk_level}</td>
                </tr>
"""

                html_content += """
            </tbody>
        </table>

        <h2>Default Rate by Segment</h2>
        <div class="chart-container">
"""

                for seg in segment_data:
                    bar_width = seg['pd_rate'] * 300  # Scale for visualization
                    html_content += f'            <div class="bar" style="width: {bar_width}px;">Segment {seg["id"]}: {seg["pd_rate"]:.2%}</div>\n'

                html_content += f"""
        </div>

        <h2>Segment Density Distribution</h2>
        <div class="chart-container">
"""

                for seg in segment_data:
                    bar_width = seg['density'] * 8  # Scale for visualization
                    html_content += f'            <div class="bar" style="width: {bar_width}px;">Segment {seg["id"]}: {seg["density"]:.1f}%</div>\n'

                html_content += f"""
        </div>

        <div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 12px;">
            Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        </div>
    </div>
</body>
</html>
"""

                with open(html_path, 'w') as f:
                    f.write(html_content)

                exported_files['html_dashboard'] = str(html_path)

                if self.verbose:
                    logger.info(f"  HTML dashboard created: {html_path}")
            except Exception as e:
                warnings.warn(f"Error creating HTML dashboard: {e}")

        if self.verbose:
            logger.info(f"\n[OK] Exported {len(exported_files)} files:")
            for output_type, path in exported_files.items():
                logger.info(f"     {output_type}: {path}")

        return exported_files

    def run_all(self, pause_for_edits: bool = True) -> None:
        """
        Run entire pipeline from start to finish.

        Args:
            pause_for_edits: If True, pause after template export for user edits
        """
        if self.verbose:
            logger.info("=" * 70)
            logger.info("RUNNING COMPLETE PIPELINE")
            logger.info("=" * 70)

        # Stage 1: Load data
        self.load_data()

        # Stage 2: Fit baseline
        self.fit_baseline()

        # Stage 3: Export template
        template_path = self.export_template()

        if pause_for_edits:
            if self.verbose:
                logger.info("=" * 70)
                logger.info("PAUSED FOR TEMPLATE EDITING")
                logger.info("=" * 70)
                logger.info(f"\n1. Edit the template: {template_path}")
                logger.info(f"2. Save your changes")
                logger.info(f"3. Run: pipeline.apply_modifications('{template_path}')")
                logger.info(f"4. Then run: pipeline.compare_results()")
                logger.info(f"5. Finally: pipeline.export_all()")
            return

        # If not pausing, skip modifications and just export
        self.export_all()

        if self.verbose:
            logger.info("=" * 70)
            logger.info("PIPELINE COMPLETE")
            logger.info("=" * 70)

    def get_state(self) -> Dict[str, Any]:
        """
        Get current pipeline state.

        Returns:
            Dictionary with state information
        """
        return {
            'state': self.state,
            'config': {
                'name': self.config.name,
                'data_source': self.config.data.source,
                'data_type': self.config.data.data_type
            },
            'data': {
                'loaded': self.state['data_loaded'],
                'train_size': len(self.X_train) if self.X_train is not None else 0,
                'val_size': len(self.X_val) if self.X_val is not None else 0,
                'n_features': len(self.feature_names) if self.feature_names else 0
            },
            'baseline': {
                'fitted': self.state['baseline_fitted'],
                'n_segments': len(np.unique(self.baseline_engine.segments_train_)) if self.baseline_engine else 0
            },
            'modified': {
                'applied': self.state['modifications_applied'],
                'n_segments': len(np.unique(self.modified_engine.segments_train_)) if self.modified_engine else 0
            }
        }

    def _generate_markdown_report(
        self,
        output_path: Path,
        segment_descriptions: Dict[int, Dict],
        viz_outputs: Dict[str, str]
    ) -> None:
        """
        Generate comprehensive markdown report with segment descriptions and visualizations.

        Args:
            output_path: Path to save markdown file
            segment_descriptions: Dictionary of segment descriptions from engine
            viz_outputs: Dictionary of visualization file paths
        """
        with open(output_path, 'w') as f:
            # Header
            f.write("# IRB Segmentation Analysis Report\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write("---\n\n")

            # Executive Summary
            f.write("## Executive Summary\n\n")
            total_obs = len(self.y_train)
            total_defaults = int(np.sum(self.y_train))
            overall_dr = float(np.mean(self.y_train))
            n_segments = len(segment_descriptions)

            f.write(f"- **Total Observations:** {total_obs:,}\n")
            f.write(f"- **Total Defaults:** {total_defaults:,}\n")
            f.write(f"- **Overall Default Rate:** {overall_dr:.2%}\n")
            f.write(f"- **Number of Segments:** {n_segments}\n\n")

            # Risk classification summary
            f.write("### Risk Distribution\n\n")
            f.write("| Risk Level | Segments | Total Observations | Default Rate |\n")
            f.write("|-----------|----------|-------------------|-------------|\n")

            risk_levels = {'Low': [], 'Medium': [], 'High': []}
            for seg_id, desc in segment_descriptions.items():
                dr = desc['statistics']['default_rate']
                if dr < 0.10:
                    risk_levels['Low'].append(desc)
                elif dr < 0.20:
                    risk_levels['Medium'].append(desc)
                else:
                    risk_levels['High'].append(desc)

            for risk, segments in risk_levels.items():
                if segments:
                    n_seg = len(segments)
                    total_seg_obs = sum(s['statistics']['n_observations'] for s in segments)
                    avg_dr = np.mean([s['statistics']['default_rate'] for s in segments])
                    f.write(f"| {risk} | {n_seg} | {total_seg_obs:,} | {avg_dr:.2%} |\n")

            f.write("\n")

            # Overall visualizations
            f.write("## Overall Segment Visualizations\n\n")

            # Feature distributions
            if 'distributions' in viz_outputs:
                f.write("### Feature Distributions by Segment\n\n")
                rel_path = Path(viz_outputs['distributions']).name
                f.write(f"![Feature Distributions]({rel_path})\n\n")
                f.write("*Shows how key features distribute within each segment, revealing segment characteristics.*\n\n")

            # Box plots
            if 'boxplots' in viz_outputs:
                f.write("### Feature Comparison (Box Plots)\n\n")
                rel_path = Path(viz_outputs['boxplots']).name
                f.write(f"![Box Plots]({rel_path})\n\n")
                f.write("*Side-by-side comparison of feature ranges across segments.*\n\n")

            # Radar chart
            if 'radar' in viz_outputs:
                f.write("### Segment Profile Comparison (Radar Chart)\n\n")
                rel_path = Path(viz_outputs['radar']).name
                f.write(f"![Radar Chart]({rel_path})\n\n")
                f.write("*Visual fingerprint of each segment's normalized feature values.*\n\n")

            # Scatter plots
            if 'scatter' in viz_outputs:
                f.write("### Feature Space Visualization\n\n")
                rel_path = Path(viz_outputs['scatter']).name
                f.write(f"![2D Scatter]({rel_path})\n\n")
                f.write("*2D projections showing natural clustering. Larger points indicate defaults.*\n\n")

            # Default drivers
            if 'drivers' in viz_outputs:
                f.write("### Default Rate Drivers by Segment\n\n")
                rel_path = Path(viz_outputs['drivers']).name
                f.write(f"![Default Drivers]({rel_path})\n\n")
                f.write("*Feature correlation with default risk per segment. Red = increases risk, Blue = decreases risk.*\n\n")

            f.write("---\n\n")

            # Individual segment profiles
            f.write("## Detailed Segment Profiles\n\n")

            for seg_id in sorted(segment_descriptions.keys()):
                desc = segment_descriptions[seg_id]
                stats = desc['statistics']

                # Determine risk level
                dr = stats['default_rate']
                if dr < 0.10:
                    risk_level = "Low Risk"
                    risk_emoji = ""
                elif dr < 0.20:
                    risk_level = "Medium Risk"
                    risk_emoji = ""
                else:
                    risk_level = "High Risk"
                    risk_emoji = ""

                f.write(f"### Segment {seg_id}: {risk_level} {risk_emoji}\n\n")

                # Statistics
                f.write("#### Statistics\n\n")
                f.write(f"- **Observations:** {stats['n_observations']:,} ({stats['density']:.2%} of population)\n")
                f.write(f"- **Defaults:** {stats['n_defaults']:,}\n")
                f.write(f"- **Default Rate:** {stats['default_rate']:.2%}\n\n")

                # Tree rule if available
                if desc['tree_rule']:
                    f.write("#### Decision Tree Rule\n\n")

                    # Check if this is a multiple-path segment (merged leaf nodes)
                    if isinstance(desc['tree_rule'], dict) and desc['tree_rule'].get('type') == 'multiple':
                        # Multiple paths merged - show all for audit purposes
                        n_paths = desc['tree_rule']['n_paths']
                        leaf_nodes = desc['tree_rule'].get('leaf_nodes', [])

                        f.write(f"**This segment was created by merging {n_paths} decision tree leaf nodes "
                               f"(nodes: {', '.join(map(str, leaf_nodes))}).**\n\n")
                        f.write("*For audit and regulatory purposes, all original decision paths are shown below:*\n\n")

                        for idx, path in enumerate(desc['tree_rule']['paths'], 1):
                            f.write(f"**Path {idx}:**\n")
                            f.write(f"```\nIF {path}\nTHEN assign to Segment {seg_id}\n```\n\n")
                    else:
                        # Single path - simple case
                        f.write(f"```\nIF {desc['tree_rule']}\nTHEN assign to Segment {seg_id}\n```\n\n")
                else:
                    f.write("#### Decision Tree Rule\n\n")
                    f.write("*Not available - segment created through post-processing adjustments (merges/splits)*\n\n")

                # Top discriminative features
                f.write("#### Key Characteristics\n\n")
                f.write("Top 5 features that distinguish this segment from the overall population:\n\n")
                f.write("| Feature | Segment Median | Population Median | Difference |\n")
                f.write("|---------|---------------|------------------|------------|\n")

                for feat_info in desc['top_discriminative_features']:
                    feat_name = feat_info['feature']
                    seg_med = feat_info['segment_median']
                    glob_med = feat_info['global_median']
                    rel_diff = feat_info['relative_difference']

                    # Format values nicely
                    if abs(seg_med) < 10:
                        seg_str = f"{seg_med:.2f}"
                        glob_str = f"{glob_med:.2f}"
                    else:
                        seg_str = f"{seg_med:,.0f}"
                        glob_str = f"{glob_med:,.0f}"

                    diff_str = f"{rel_diff:.1%}"
                    f.write(f"| {feat_name} | {seg_str} | {glob_str} | {diff_str} |\n")

                f.write("\n")

                # Feature profile table (comprehensive)
                f.write("#### Complete Feature Profile\n\n")
                f.write("<details>\n")
                f.write("<summary>Click to expand full feature statistics</summary>\n\n")
                f.write("| Feature | 5th %ile | Median | 95th %ile | Mean | Std Dev |\n")
                f.write("|---------|----------|--------|-----------|------|--------|\n")

                for feat_name, feat_profile in desc['feature_profile'].items():
                    p5 = feat_profile['p5']
                    median = feat_profile['median']
                    p95 = feat_profile['p95']
                    mean = feat_profile['mean']
                    std = feat_profile['std']

                    # Format numbers
                    if abs(median) < 10:
                        fmt_str = "{:.2f}"
                    else:
                        fmt_str = "{:,.0f}"

                    f.write(f"| {feat_name} | {fmt_str.format(p5)} | {fmt_str.format(median)} | "
                           f"{fmt_str.format(p95)} | {fmt_str.format(mean)} | {fmt_str.format(std)} |\n")

                f.write("\n</details>\n\n")

                # Adjustment history
                history = desc['adjustment_history']
                has_history = (
                    history['created_by_merges'] or
                    history['created_by_splits'] or
                    history['created_by_forced_splits'] or
                    history['has_monotonicity_violations']
                )

                if has_history:
                    f.write("#### Adjustment History\n\n")
                    f.write("*This segment was created or modified through post-processing:*\n\n")

                    if history['created_by_merges']:
                        f.write("**Merges:**\n")
                        for merge in history['created_by_merges']:
                            f.write(f"- Absorbed segment {merge['merged_segment']} "
                                   f"(reason: {merge['reason']}, {merge['original_defaults']} defaults)\n")
                        f.write("\n")

                    if history['created_by_forced_splits']:
                        f.write("**Forced Splits:**\n")
                        for split in history['created_by_forced_splits']:
                            f.write(f"- Split from segment {split['parent_segment']} "
                                   f"on {split['feature']} at threshold {split['threshold']}\n")
                        f.write("\n")

                    if history['has_monotonicity_violations']:
                        f.write("**Monotonicity Considerations:**\n")
                        for viol in history['has_monotonicity_violations']:
                            f.write(f"- {viol['feature']} ({viol['direction']}) "
                                   f"relationship with segment {viol['other_segment']}\n")
                        f.write("\n")

                f.write("---\n\n")

            # Footer
            f.write("## Notes\n\n")
            f.write("### PDF Conversion\n\n")
            f.write("To convert this report to PDF, run:\n\n")
            f.write("```bash\n")
            f.write("pandoc segment_report.md -o segment_report.pdf --pdf-engine=xelatex\n")
            f.write("```\n\n")
            f.write("### File Locations\n\n")
            f.write("All visualization files are saved in the same directory as this report.\n\n")

        if self.verbose:
            logger.info(f"    Markdown report generated: {output_path}")
